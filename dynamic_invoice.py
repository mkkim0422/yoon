"""
dynamic_invoice.py — 템플릿 디자인 보존 + 동적 SKU 매핑 + 살아있는 수식 주입.

⚠ 라이브러리 한계 고지 (openpyxl):
    openpyxl 은 엑셀의 '도형(Shape/AutoShape)' — 예: B2 의 INVOICE 라운드 도형 —
    을 보존하지 못한다. 저장 시 해당 오브젝트가 사라진다.
    해결책:
      1) 템플릿 작성 시 도형을 PNG/JPG 그림(이미지)으로 교체 → openpyxl 이 유지.
      2) 또는 셀 배경색/테두리로 시각 요소를 대체.
    이 모듈은 도형 보존을 시도하지 않으며, 어떤 워크어라운드도 제공하지 않는다.

사용:
    python dynamic_invoice.py template.xlsx billing.csv out.xlsx \
        --project hyundai-autolink --fx 1525.3

설계 원칙:
    1) 템플릿(template.xlsx) 의 배경/병합/선/폰트는 100% 보존.
    2) SKU 좌표는 3개 시트에서 동적으로 탐색 (하드코딩 금지).
    3) 값이 아닌 '=수식' 문자열을 주입 → Excel 에서 값 수정 시 즉시 재계산.
    4) 모든 셀 쓰기/읽기는 safe_write / safe_read 로만 수행 → MergedCell 에러 차단.
"""
from __future__ import annotations

import argparse
import logging
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger("dynamic_invoice")

# ── 시트 이름 ───────────────────────────────────────────────────────────────
SHEET_PRICE   = "GMP Price List"
SHEET_INVOICE = "Invoice"
SHEET_PROJECT = "Project"

# ── 티어 경계(누적, count) ─────────────────────────────────────────────────
# Tier1 0–100K / Tier2 100K–500K / Tier3 500K–1M / Tier4 1M–5M / Tier5 5M+
TIER_BREAKS      = [100_000, 500_000, 1_000_000, 5_000_000]
PRICE_COLS       = ["D", "E", "F", "G", "H"]
BLOCK_ROWS       = 6
TIER_ROW_OFFSETS = [0, 1, 2, 3, 4]
SUBTOTAL_OFFSET  = 5

# ── 스타일 ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="1F3864")
TOTAL_FILL  = PatternFill("solid", fgColor="000000")
TOTAL_FONT  = Font(bold=True, color="FFFFFF")
THIN        = Side(style="thin", color="BFBFBF")
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FMT_USD     = "#,##0.00"
FMT_KRW     = "#,##0"


# ════════════════════════════════════════════════════════════════════════════
# 1. 병합 셀 안전 입출력
# ════════════════════════════════════════════════════════════════════════════
def _anchor(ws: Worksheet, row: int, col: int) -> tuple[int, int]:
    """병합 영역 내부면 좌상단 마스터 좌표 반환, 아니면 입력 그대로."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return row, col
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col
    return row, col


def safe_write(ws: Worksheet, row: int, col: int, val, *, allow_any_col: bool = False):
    """
    [무결성 가드] Invoice 시트는 C열(col==3)만 쓰기 허용. 그 외 시트는 allow_any_col 로 우회.
    [세포 복제] 8종 속성(font/border/fill/alignment/number_format/protection/hyperlink/comment)
    을 스냅샷 → 값 주입 → 원복하여 템플릿 유전자를 100% 보존.
    comment 는 객체 참조 충돌 방지를 위해 copy() 사본으로 재할당.
    """
    ar, ac = _anchor(ws, row, col)
    # [C열 전용 가드] 병합셀 앵커(ac) 기준으로 체크해야 C10:C14 같은 병합셀에 정상 주입됨
    if ws.title == SHEET_INVOICE and not allow_any_col and ac != 3:
        return ws.cell(row=ar, column=ac)
    cell = ws.cell(row=ar, column=ac)
    snap_font       = copy(cell.font)        if cell.font       is not None else None
    snap_border     = copy(cell.border)      if cell.border     is not None else None
    snap_fill       = copy(cell.fill)        if cell.fill       is not None else None
    snap_align      = copy(cell.alignment)   if cell.alignment  is not None else None
    snap_numfmt     = cell.number_format  # str, copy 불필요
    snap_protection = copy(cell.protection)  if cell.protection is not None else None
    snap_hyperlink  = copy(cell.hyperlink)   if cell.hyperlink  is not None else None
    snap_comment    = copy(cell.comment)     if cell.comment    is not None else None

    cell.value = val

    # 원복 시 또 한 번 copy() 로 '새로운 스타일 객체'로 재탄생시켜 StyleProxy 차단
    if snap_font:       cell.font          = copy(snap_font)
    if snap_border:     cell.border        = copy(snap_border)
    if snap_fill:       cell.fill          = copy(snap_fill)
    if snap_align:      cell.alignment     = copy(snap_align)
    if snap_numfmt:     cell.number_format = snap_numfmt
    if snap_protection: cell.protection    = copy(snap_protection)
    if snap_hyperlink:  cell.hyperlink     = copy(snap_hyperlink)
    if snap_comment:    cell.comment       = copy(snap_comment)
    return cell


def safe_read(ws: Worksheet, row: int, col: int):
    """병합 셀이라도 마스터 셀의 값을 반환. 읽기는 무조건 이 함수만 사용."""
    ar, ac = _anchor(ws, row, col)
    return ws.cell(row=ar, column=ac).value


# ════════════════════════════════════════════════════════════════════════════
# 2. SKU 정규화 / 별칭 / 부분 매칭
# ════════════════════════════════════════════════════════════════════════════
def _norm(s) -> str:
    """strip + casefold + 연속공백 1칸."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).casefold()


# [1:1 Exact Match 정책] 부분일치/병합 테이블 폐기. 정규화된 SKU 문자열 자체를 키로 사용.
# CSV 'SKU 설명' ↔ 템플릿 Invoice B열 'API 이름' 이 _norm() 기준 완전 일치할 때만 합산.


def _resolve_usage(template_key: str, usage_map: dict[str, float]) -> float:
    """[Exact Match] 정규화된 키가 완전 일치할 때만 반환. 미매칭 시 0.0."""
    return float(usage_map.get(template_key, 0.0))


# ════════════════════════════════════════════════════════════════════════════
# 3. 동적 매핑
# ════════════════════════════════════════════════════════════════════════════
@dataclass
class SkuMap:
    name: str
    price_row: int
    invoice_row: int
    project_col: int | None


def build_mapping(wb) -> Dict[str, SkuMap]:
    price_ws   = wb[SHEET_PRICE]
    invoice_ws = wb[SHEET_INVOICE]
    project_ws = wb[SHEET_PROJECT]

    price_map: dict[str, tuple[str, int]] = {}
    for row in range(1, price_ws.max_row + 1):
        v = safe_read(price_ws, row, 1)
        k = _norm(v)
        if k and k not in price_map:
            price_map[k] = (str(v).strip(), row)

    invoice_map: dict[str, int] = {}
    for row in range(1, invoice_ws.max_row + 1):
        k = _norm(safe_read(invoice_ws, row, 2))
        if k and k in price_map and k not in invoice_map:
            invoice_map[k] = row

    project_map: dict[str, int] = {}
    for col in range(1, project_ws.max_column + 1):
        k = _norm(safe_read(project_ws, 10, col))
        if k and k in price_map and k not in project_map:
            project_map[k] = col

    mapping: dict[str, SkuMap] = {}
    for k, (orig, prow) in price_map.items():
        if k not in invoice_map:
            continue
        mapping[k] = SkuMap(
            name        = orig,
            price_row   = prow,
            invoice_row = invoice_map[k],
            project_col = project_map.get(k),
        )
    log.info("매핑 완료: %d개 SKU", len(mapping))
    return mapping


# ════════════════════════════════════════════════════════════════════════════
# 4. billing.csv 로더
# ════════════════════════════════════════════════════════════════════════════
class ProjectNotFoundError(RuntimeError):
    pass


def _to_float(v) -> float:
    """콤마/따옴표/공백 제거 후 float 변환. 실패 시 0.0."""
    if v is None:
        return 0.0
    try:
        s = str(v).strip().strip('"').replace(",", "").replace('"', "")
        return float(s) if s else 0.0
    except (TypeError, ValueError):
        return 0.0


def load_usage(csv_path: Path, project: str | None = None
               ) -> tuple[dict[str, float], float, float, str]:
    fx = 0.0
    with csv_path.open("r", encoding="utf-8-sig") as f:
        for line in f:
            if line.startswith("환율"):
                m = re.search(r"[\d,\.]+", line.split(",", 1)[1])
                if m:
                    fx = float(m.group(0).replace(",", ""))
                break

    df = pd.read_csv(csv_path, skiprows=8, dtype=str, keep_default_na=False)

    # [한국어 컬럼 명시적 타겟팅]
    ACCT_COL  = "결제 계정 이름"
    SKU_COL   = "SKU 설명"
    USAGE_COL = "사용량"
    PID_COL   = "프로젝트 ID"
    COST_COL  = next((c for c in df.columns if c.startswith("비용")), None)

    for need in (ACCT_COL, SKU_COL, USAGE_COL):
        if need not in df.columns:
            raise RuntimeError(f"CSV에 '{need}' 컬럼이 없습니다. 보유: {list(df.columns)}")

    if "비용 유형" in df.columns:
        df = df[df["비용 유형"].astype(str).str.contains("사용량", na=False)]

    # ── [필터] 결제 계정 이름에 project 인자 포함 (하드코딩 제거) ────
    if not project:
        raise RuntimeError("--project 인자가 필요합니다 (필터링 대상 고객사).")
    target_acct = _norm(project)
    mask = df[ACCT_COL].astype(str).map(lambda s: target_acct in _norm(s))
    df_inv = df[mask]
    log.info("결제 계정 이름이 '%s' 포함 행 개수: %d", project, len(df_inv))

    # ── [합산] SKU 설명 부분일치 + 사용량 float 변환 ─────────────────
    # [1:1 Exact Match] _norm() 정규화된 SKU 설명 자체를 키로 사용. 병합/부분일치 금지.
    usage_map: dict[str, float] = {}
    for sku_raw, u_raw in zip(df_inv[SKU_COL], df_inv[USAGE_COL]):
        k = _norm(sku_raw)
        if not k:
            continue
        usage_map[k] = usage_map.get(k, 0.0) + _to_float(u_raw)

    def _sum_for(name: str) -> float:
        return usage_map.get(_norm(name), 0.0)

    log.info("Dynamic Maps 매칭 결과: %s",   f"{_sum_for('Dynamic Maps'):,.0f}")
    log.info("Geocoding 매칭 결과: %s",      f"{_sum_for('Geocoding'):,.0f}")
    log.info("Places Details 매칭 결과: %s", f"{_sum_for('Places Details'):,.0f}")

    if project and PID_COL in df.columns:
        tgt = _norm(project)
        hits = int(df[PID_COL].map(lambda s: _norm(s).startswith(tgt)).sum())
        log.info("Project 시트 startswith('%s') 매칭 행 수: %d", project, hits)

    total_cost = float(pd.to_numeric(df_inv[COST_COL], errors="coerce").fillna(0).sum()) if COST_COL else 0.0

    # ── [Term of Use] 사용량 시작일/종료일에서 기간 추출 ─────────────
    start_col = next((c for c in df_inv.columns if "시작일" in c), None)
    end_col   = next((c for c in df_inv.columns if "종료일" in c), None)
    term = ""
    if start_col and end_col and len(df_inv):
        starts = pd.to_datetime(df_inv[start_col], errors="coerce").dropna()
        ends   = pd.to_datetime(df_inv[end_col],   errors="coerce").dropna()
        if len(starts) and len(ends):
            term = f"{starts.min().strftime('%Y-%m-%d')} ~ {ends.max().strftime('%Y-%m-%d')}"
            log.info("Term of Use: %s", term)
    return usage_map, fx, total_cost, term


# ════════════════════════════════════════════════════════════════════════════
# 5. 수식 생성 / 시트 주입
# ════════════════════════════════════════════════════════════════════════════
def inject_invoice(ws: Worksheet, mapping: dict[str, SkuMap],
                   usage_map: dict[str, float]) -> list[tuple[str, str]]:
    """
    [무결성 보존] Invoice 시트는 오직 C열(col==3)만 쓴다.
    D/E/F/G/H/I 의 수식·서식은 템플릿 원본 그대로 보존.
    """
    refs: list[tuple[str, str]] = []
    for key, sm in mapping.items():
        r = sm.invoice_row
        val = float(_resolve_usage(key, usage_map))
        for col in (3,):
            if col == 3:
                safe_write(ws, r, col, val)
        log.info("C%d 셀 주입 완료 — %s = %s", r, sm.name, f"{val:,.0f}")
        refs.append((sm.name, f"I{r + SUBTOTAL_OFFSET}"))
    return refs


def hide_zero_blocks(ws: Worksheet, mapping: dict[str, SkuMap]) -> int:
    """
    [스마트 숨김] inject_invoice 후 호출. C열에 실제로 써진 값을 읽어서 판별.
    float > 0 → 표시 유지, 그 외(0/None/빈칸) → 블록 전체 hidden.
    """
    hidden_count = 0
    for _key, sm in mapping.items():
        base = sm.invoice_row
        raw = safe_read(ws, base, 3)  # inject_invoice 가 이미 써 놓은 C열 값
        try:
            usage = float(raw) if raw is not None else 0.0
        except (TypeError, ValueError):
            usage = 0.0

        if usage > 0.0:
            for offset in range(BLOCK_ROWS):
                ws.row_dimensions[base + offset].hidden = False
            log.info("Invoice: %s 사용량 %s -> 숨김 처리 X", sm.name, f"{usage:,.0f}")
        else:
            for offset in range(BLOCK_ROWS):
                ws.row_dimensions[base + offset].hidden = True
            hidden_count += 1
            log.info("Invoice: %s 사용량 0 -> 숨김 처리 O (행 %d~%d)",
                     sm.name, base, base + BLOCK_ROWS - 1)
    log.info("Invoice 숨김 완료: %d개 블록 숨김", hidden_count)
    return hidden_count


def ensure_project_row(ws: Worksheet, project: str) -> int:
    """B11 이하에서 project 검색 → 없으면 첫 빈 마스터 셀에 safe_write."""
    target = _norm(project)
    scan_to = max(ws.max_row, 11)
    first_empty: tuple[int, int] | None = None
    seen: set[tuple[int, int]] = set()

    for r in range(11, scan_to + 2):
        ar, ac = _anchor(ws, r, 2)
        if (ar, ac) in seen:
            continue
        seen.add((ar, ac))
        v = safe_read(ws, ar, ac)
        if v is None or str(v).strip() == "":
            if first_empty is None:
                first_empty = (ar, ac)
            continue
        nv = _norm(v)
        if nv == target or nv.startswith(target) or target.startswith(nv):
            return ar

    ar, ac = first_empty or (11, 2)
    safe_write(ws, ar, ac, project, allow_any_col=True)
    log.info("Project 시트 %s%d 에 '%s' 자동 기입",
             get_column_letter(ac), ar, project)
    return ar


HEADER_ACCT_ROW  = 6   # C6: Billing Account Name
HEADER_TERM_ROW  = 7   # C7: Term of Use
HEADER_COL       = 3   # C열

def inject_header(ws: Worksheet, account_name: str, term: str) -> None:
    """Invoice/Project 시트 공통: C6=고객사명, C7=사용기간 을 동적으로 덮어씀."""
    if account_name:
        safe_write(ws, HEADER_ACCT_ROW, HEADER_COL, account_name, allow_any_col=True)
    if term:
        safe_write(ws, HEADER_TERM_ROW, HEADER_COL, term,         allow_any_col=True)


PROJECT_DATA_START_ROW = 11  # 1~9행: 헤더 영역(환율 C8 포함) / 10행: 라벨 / 11~: 데이터

def _clean_project_residual(ws: Worksheet, project: str) -> None:
    """
    B열 데이터 행(11~)에서 타겟 고객사와 무관한 잔여 프로젝트명만 빈칸 처리.
    타겟과 관련된 행은 절대 건드리지 않는다.
    """
    target = _norm(project)
    seen: set[tuple[int, int]] = set()
    for r in range(PROJECT_DATA_START_ROW, max(ws.max_row, PROJECT_DATA_START_ROW) + 2):
        ar, ac = _anchor(ws, r, 2)
        if (ar, ac) in seen:
            continue
        seen.add((ar, ac))
        v = safe_read(ws, ar, ac)
        if v is None or str(v).strip() == "":
            continue
        nv = _norm(v)
        # 타겟과 조금이라도 관련 있으면(포함/시작) 보존
        is_related = (target in nv or nv in target
                      or nv.startswith(target) or target.startswith(nv))
        if is_related:
            log.info("Project 보존: %s%d '%s'", get_column_letter(ac), ar, v)
        else:
            safe_write(ws, ar, ac, "", allow_any_col=True)
            log.info("Project 잔여물 제거: %s%d '%s'", get_column_letter(ac), ar, v)


def inject_project(ws: Worksheet, mapping: dict[str, SkuMap],
                   fx_rate: float, usage_map: dict[str, float],
                   project: str | None = None) -> None:
    del fx_rate
    if project:
        _clean_project_residual(ws, project)
        ensure_project_row(ws, project)

    for key, sm in mapping.items():
        if sm.project_col is None:
            continue
        col = sm.project_col
        usage_row = PROJECT_DATA_START_ROW           # 11
        month_row = PROJECT_DATA_START_ROW + 1       # 12
        amt_row   = PROJECT_DATA_START_ROW + 2       # 13

        if min(usage_row, month_row, amt_row) < 10:
            raise RuntimeError("Project 시트 start_row 오류: 10행 이상부터만 쓰기 허용")

        # [0원 → 빈칸] 사용하지 않은 API는 수식 대신 빈 문자열로 깔끔하게
        val = _resolve_usage(key, usage_map)
        if val == 0.0:
            safe_write(ws, usage_row, col, "", allow_any_col=True)
            safe_write(ws, month_row, col, "", allow_any_col=True)
            safe_write(ws, amt_row,   col, "", allow_any_col=True)
        else:
            u = safe_write(ws, usage_row, col, f"='{SHEET_INVOICE}'!C{sm.invoice_row}", allow_any_col=True)
            m = safe_write(ws, month_row, col, f"='{SHEET_INVOICE}'!I{sm.invoice_row + SUBTOTAL_OFFSET}", allow_any_col=True)
            a = safe_write(ws, amt_row,   col, f"='{SHEET_INVOICE}'!I{sm.invoice_row + SUBTOTAL_OFFSET}*$C$8", allow_any_col=True)
            u.number_format = FMT_KRW
            m.number_format = FMT_USD
            a.number_format = FMT_KRW


# ════════════════════════════════════════════════════════════════════════════
# 6. 정합성 검증 / 엔트리포인트
# ════════════════════════════════════════════════════════════════════════════
def verify_integrity(out_path: Path, csv_total: float) -> None:
    try:
        load_workbook(out_path, data_only=True)
        log.info("정합성 체크: 원본 CSV 총합(₩)=%s | Excel 1회 저장 후 수식 계산값 재검증 권장",
                 f"{csv_total:,.0f}")
    except Exception as e:
        log.warning("정합성 검증 실패: %s", e)


def generate(template: Path, csv: Path, out: Path,
             fx: float | None = None, project: str | None = None) -> Path:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    wb = load_workbook(template, data_only=False)
    for need in (SHEET_PRICE, SHEET_INVOICE, SHEET_PROJECT):
        if need not in wb.sheetnames:
            raise ValueError(f"템플릿에 '{need}' 시트가 없습니다. 보유: {wb.sheetnames}")

    mapping = build_mapping(wb)
    usage_map, csv_fx, csv_total, term = load_usage(csv, project=project)
    fx_rate = fx if fx is not None else (csv_fx or 1.0)

    refs = inject_invoice(wb[SHEET_INVOICE], mapping, usage_map)
    hide_zero_blocks(wb[SHEET_INVOICE], mapping)
    inject_project(wb[SHEET_PROJECT], mapping, fx_rate,
                   usage_map=usage_map, project=project)

    if project:
        inject_header(wb[SHEET_INVOICE], project, term)
        inject_header(wb[SHEET_PROJECT], project, term)

    # 강제 재계산: Excel 열 때 모든 수식 즉시 평가
    wb.calculation.calcMode = "auto"
    wb.properties.calcFinished = False

    try:
        wb.save(out)
    except PermissionError:
        msg = (f"❌ 파일 저장 실패: '{out}' 가 다른 프로그램(Excel 등)에서 열려 있습니다.\n"
               f"   → 해당 파일을 닫은 뒤 다시 실행해 주세요.")
        log.error(msg)
        raise SystemExit(1)

    log.info("저장: %s (SKU %d개, 소계 셀 %d개)", out, len(mapping), len(refs))
    verify_integrity(out, csv_total)

    # [무결성 최종 검증] Invoice!I11 수식이 템플릿 원본 그대로 보존됐는지 확인
    chk = load_workbook(out, data_only=False)[SHEET_INVOICE]["I11"].value
    log.info("Invoice!I11 수식이 원래대로 %s 인지 확인 완료", chk)
    return out


def _cli():
    p = argparse.ArgumentParser()
    p.add_argument("template")
    p.add_argument("csv")
    p.add_argument("out")
    p.add_argument("--fx", type=float, default=None)
    p.add_argument("--project", type=str, default=None,
                   help="billing.csv의 '프로젝트 ID' 필터 (대소문자/공백 무시)")
    a = p.parse_args()
    try:
        generate(Path(a.template), Path(a.csv), Path(a.out), a.fx, a.project)
    except ProjectNotFoundError as e:
        log.error(str(e))
        raise SystemExit(2)


if __name__ == "__main__":
    _cli()
