"""
통합 테스트 - Mock 데이터로 전체 파이프라인 실행 후 test_invoice_raw.xlsx 생성 검증

실행: python -m pytest tests/test_integration.py -v
     또는 단독 실행: python tests/test_integration.py
"""
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from billing.engine import calculate_billing
from billing.loader import load_sku_master, load_usage_rows
from billing.models import BillingLineItem


# ── 1. Mock 데이터 ─────────────────────────────────────────────────────────

# 맵핑 마스터: skus + sku_tiers JOIN 결과를 흉내 낸 딕셔너리 리스트
MOCK_SKU_MASTER_ROWS = [
    # Dynamic Maps  (무료 1만, 1구간 100K @$7, 2구간 @$5.6)
    {
        "sku_id": "maps.dynamic",
        "sku_name": "Dynamic Maps",
        "is_billable": True,
        "category": "Maps",
        "free_usage_cap": 10_000,
        "tier_number": 1,
        "tier_limit": 100_000,   # 누적 상한 100K
        "tier_cpm": 7.0,
    },
    {
        "sku_id": "maps.dynamic",
        "sku_name": "Dynamic Maps",
        "is_billable": True,
        "category": "Maps",
        "free_usage_cap": 10_000,
        "tier_number": 2,
        "tier_limit": None,      # 마지막 구간 (500K 초과도 동일 단가)
        "tier_cpm": 5.6,
    },
    # Places Details  (무료 5천, 1구간 100K @$17, 2구간 @$13.6)
    {
        "sku_id": "places.details",
        "sku_name": "Places Details",
        "is_billable": True,
        "category": "Places",
        "free_usage_cap": 5_000,
        "tier_number": 1,
        "tier_limit": 100_000,
        "tier_cpm": 17.0,
    },
    {
        "sku_id": "places.details",
        "sku_name": "Places Details",
        "is_billable": True,
        "category": "Places",
        "free_usage_cap": 5_000,
        "tier_number": 2,
        "tier_limit": None,
        "tier_cpm": 13.6,
    },
    # 세금  (비과금)
    {
        "sku_id": "tax.vat",
        "sku_name": "세금",
        "is_billable": False,
        "category": "Tax",
        "free_usage_cap": 0,
        "tier_number": None,
        "tier_limit": None,
        "tier_cpm": None,
    },
]

# 로우 데이터: usage_raw 테이블을 흉내 낸 딕셔너리 리스트
MOCK_USAGE_RAW_ROWS = [
    # popo-01: Dynamic Maps 15만 건
    {"billing_month": "2025-03", "project_id": "popo-01", "sku_id": "maps.dynamic",    "usage_amount": 150_000},
    # popo-01: 세금 1건 (비과금 → 필터 제외 대상)
    {"billing_month": "2025-03", "project_id": "popo-01", "sku_id": "tax.vat",          "usage_amount": 1},
    # popo-02: Places Details 3천 건 (무료 5천 이하 → 청구액 0)
    {"billing_month": "2025-03", "project_id": "popo-02", "sku_id": "places.details",   "usage_amount": 3_000},
]

EXCHANGE_RATE = Decimal("1350")
MARGIN_RATE   = Decimal("1.12")
OUTPUT_PATH   = Path(__file__).parent.parent / "test_invoice_raw.xlsx"


# ── 2. Excel 내보내기 ──────────────────────────────────────────────────────

def export_to_excel(items: list[BillingLineItem], output_path: Path) -> None:
    """BillingLineItem 리스트를 xlsx 파일로 저장."""
    wb = openpyxl.Workbook()

    # ── 시트 1: 인보이스 요약 ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Invoice Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center")

    summary_headers = [
        "정산월", "프로젝트", "SKU ID", "SKU명",
        "총사용량", "무료차감", "청구대상",
        "소계(USD)", "환율", "마진율", "최종(KRW)",
    ]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for item in items:
        ws_summary.append([
            item.billing_month,
            item.project_id,
            item.sku_id,
            item.sku_name,
            item.total_usage,
            item.free_cap_applied,
            item.billable_usage,
            float(item.subtotal_usd),
            float(item.exchange_rate),
            float(item.margin_rate),
            float(item.final_krw),
        ])

    # 열 너비 자동 조정
    col_widths = [10, 12, 18, 18, 12, 12, 12, 12, 8, 8, 14]
    for i, width in enumerate(col_widths, start=1):
        ws_summary.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = width

    # ── 시트 2: 구간별 세부 내역 ──────────────────────────────────────────
    ws_detail = wb.create_sheet("Tier Breakdown")

    detail_headers = [
        "프로젝트", "SKU ID", "구간번호",
        "구간사용량", "구간단가(CPM)", "구간금액(USD)",
    ]
    ws_detail.append(detail_headers)
    for cell in ws_detail[1]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    for item in items:
        for tb in item.tier_breakdown:
            ws_detail.append([
                item.project_id,
                item.sku_id,
                tb.tier_number,
                tb.usage_in_tier,
                float(tb.tier_cpm),
                float(tb.amount_usd),
            ])

    for i, width in enumerate([12, 18, 10, 12, 14, 14], start=1):
        ws_detail.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = width

    wb.save(output_path)


# ── 3. 파이프라인 실행 함수 ────────────────────────────────────────────────

def run_pipeline() -> list[BillingLineItem]:
    sku_master  = load_sku_master(MOCK_SKU_MASTER_ROWS)
    usage_rows  = load_usage_rows(MOCK_USAGE_RAW_ROWS)
    line_items  = calculate_billing(usage_rows, sku_master, EXCHANGE_RATE, MARGIN_RATE)
    export_to_excel(line_items, OUTPUT_PATH)
    return line_items


# ── 4. 단언(assert) 테스트 ─────────────────────────────────────────────────

class TestIntegrationPipeline:
    """Mock 데이터로 전체 파이프라인을 실행하고 결과를 검증한다."""

    def setup_method(self):
        self.items = run_pipeline()
        # project_id + sku_id → BillingLineItem 인덱스
        self.idx = {(i.project_id, i.sku_id): i for i in self.items}

    # ── 결과 건수 ────────────────────────────────────────────────────────
    def test_result_count(self):
        """세금(비과금) 제외 → popo-01/maps.dynamic + popo-02/places.details = 2건"""
        assert len(self.items) == 2

    def test_tax_excluded(self):
        """is_billable=False인 세금 SKU는 결과에 없어야 한다"""
        assert ("popo-01", "tax.vat") not in self.idx

    # ── popo-01 / Dynamic Maps ──────────────────────────────────────────
    def test_popo01_total_usage(self):
        item = self.idx[("popo-01", "maps.dynamic")]
        assert item.total_usage == 150_000

    def test_popo01_free_cap(self):
        item = self.idx[("popo-01", "maps.dynamic")]
        assert item.free_cap_applied == 10_000
        assert item.billable_usage   == 140_000

    def test_popo01_tier1_breakdown(self):
        """T1: 100,000건 × $7/1000 = $700"""
        item = self.idx[("popo-01", "maps.dynamic")]
        t1 = item.tier_breakdown[0]
        assert t1.tier_number   == 1
        assert t1.usage_in_tier == 100_000
        assert t1.tier_cpm      == Decimal("7.0")
        assert t1.amount_usd    == Decimal("700")

    def test_popo01_tier2_breakdown(self):
        """T2: 40,000건 × $5.6/1000 = $224"""
        item = self.idx[("popo-01", "maps.dynamic")]
        t2 = item.tier_breakdown[1]
        assert t2.tier_number   == 2
        assert t2.usage_in_tier == 40_000
        assert t2.tier_cpm      == Decimal("5.6")
        assert t2.amount_usd    == Decimal("224")

    def test_popo01_subtotal_usd(self):
        """$700 + $224 = $924"""
        item = self.idx[("popo-01", "maps.dynamic")]
        assert item.subtotal_usd == Decimal("924")

    def test_popo01_final_krw(self):
        """$924 × 1350 × 1.12 = 1,397,088원"""
        item = self.idx[("popo-01", "maps.dynamic")]
        expected = (Decimal("924") * EXCHANGE_RATE * MARGIN_RATE).quantize(Decimal("1"))
        assert item.final_krw == expected  # 1,397,088

    # ── popo-02 / Places Details ────────────────────────────────────────
    def test_popo02_total_usage(self):
        item = self.idx[("popo-02", "places.details")]
        assert item.total_usage == 3_000

    def test_popo02_within_free_cap(self):
        """3,000건 < 무료 5,000건 → 청구 0"""
        item = self.idx[("popo-02", "places.details")]
        assert item.free_cap_applied == 3_000
        assert item.billable_usage   == 0
        assert item.subtotal_usd     == Decimal("0")
        assert item.final_krw        == Decimal("0")

    def test_popo02_no_tier_breakdown(self):
        """청구 대상 없으면 구간 내역도 없다"""
        item = self.idx[("popo-02", "places.details")]
        assert item.tier_breakdown == []

    # ── Excel 파일 검증 ─────────────────────────────────────────────────
    def test_excel_file_created(self):
        """test_invoice_raw.xlsx 파일이 생성되어야 한다"""
        assert OUTPUT_PATH.exists(), f"파일 없음: {OUTPUT_PATH}"
        assert OUTPUT_PATH.stat().st_size > 0

    def test_excel_has_two_sheets(self):
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        assert "Invoice Summary" in wb.sheetnames
        assert "Tier Breakdown"  in wb.sheetnames

    def test_excel_summary_row_count(self):
        """헤더 1행 + 데이터 2행 = 총 3행"""
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        ws = wb["Invoice Summary"]
        assert ws.max_row == 3

    def test_excel_popo01_krw_value(self):
        """Summary 시트에서 popo-01 행의 최종 KRW 값 확인"""
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        ws = wb["Invoice Summary"]
        # 헤더 제외, 데이터 행 순회
        krw_col = 11  # 'final_krw' 열 인덱스
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "popo-01":   # project_id 열
                assert row[krw_col - 1] == float(
                    (Decimal("924") * EXCHANGE_RATE * MARGIN_RATE).quantize(Decimal("1"))
                )
                found = True
        assert found, "popo-01 행을 Excel에서 찾지 못했습니다"


# ── 5. 단독 실행 ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    items = run_pipeline()

    print(f"\n{'='*60}")
    print(f"  파이프라인 실행 완료 - 결과 {len(items)}건")
    print(f"{'='*60}")
    for item in items:
        print(f"\n[{item.project_id}] {item.sku_name}")
        print(f"  총사용량     : {item.total_usage:>10,} 건")
        print(f"  무료차감     : {item.free_cap_applied:>10,} 건")
        print(f"  청구대상     : {item.billable_usage:>10,} 건")
        for tb in item.tier_breakdown:
            print(f"  ├ T{tb.tier_number} {tb.usage_in_tier:>8,}건 × ${tb.tier_cpm}/1000"
                  f" = ${tb.amount_usd:>10,.2f}")
        print(f"  소계(USD)    : ${item.subtotal_usd:>10,.4f}")
        print(f"  최종(KRW)    :  {item.final_krw:>10,} 원")

    print(f"\n  Excel 저장 완료 → {OUTPUT_PATH}")
    print(f"{'='*60}\n")
