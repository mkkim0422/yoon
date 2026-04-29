"""
project_sheet.py — Project 시트 생성기 (API별 3컬럼 × 4행 격자 구조)

블록 구조 (프로젝트 1개 = 6행):
  Col A (LABEL_COL): 빈 여백
  Col B (PROJ_COL) : 프로젝트명 (세로병합 4행, WHITE)
  Col C+           : API별 데이터 (1 API당 3컬럼)

  3컬럼 구조 (per API):
    api_left  : C (price/label 좌측)
    api_mid   : D (price/label 우측, left와 병합)
    api_right : E (amount 단독 컬럼)

  6행 구조:
    행1: 서비스명  (3열 C:E 병합, #E2EFDA, 굵게)
    행2: 사용량    (3열 C:E 병합, #E2EFDA, 우측 정렬)
    행3: 라벨      (C:D 병합=monthly unit price($) | E=amount, WHITE)
    행4: 값        (C:D 병합=단가 | E=KRW 금액, WHITE)
    행5: toal($)   라벨(B:D 3칸 병합) + 값(E열)
    행6: toal(₩)  라벨(B:D 3칸 병합) + 값(E열), #E2EFDA
"""

import re
import calendar as _cal
from datetime import date as _date
from decimal import Decimal
import os

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    _HAS_IMAGE = True
except ImportError:
    _HAS_IMAGE = False

# cm → EMU 변환 상수 (1 cm = 360000 EMU)
_CM_TO_EMU = 360000


# ─────────────────────────────────────────────────────────────────────────────
# API 컬럼 표시 순서
# ─────────────────────────────────────────────────────────────────────────────
PREFERRED_API_ORDER = [
    "Dynamic Maps",
    "Geocoding",
    "Autocomplete - Per Request",
    "Autocomplete without Places Details - Per Session",
    "Query Autocomplete - Per Request",
    "Places Details",
    "Basic Data",
    "Contact Data",
    "Atmosphere Data",
    "Places - Text Search",
    "Find Place",
    "Directions",
]

API_DISPLAY_NAMES: dict[str, str] = {
    "Dynamic Maps":                                      "Dynamic Maps",
    "Basic Data":                                        "Basic Data",
    "Contact Data":                                      "Contact Data",
    "Atmosphere Data":                                   "Atmosphere\nData",
    "Find Place":                                        "Find Place",
    "Geocoding":                                         "Geocoding",
    "Autocomplete - Per Request":                        "Autocomplete\n/Pr Request",
    "Autocomplete without Places Details - Per Session": "Autocomplete\nw/o Places Details\nPer Session",
    "Places Details":                                    "Places Details",
    "Places - Text Search":                              "Places\nText Search",
    "Query Autocomplete - Per Request":                  "Query Autocomplete\nPer Request",
    "Directions":                                        "Directions",
}

# ─────────────────────────────────────────────────────────────────────────────
# FORCED unit prices — 비워둠.
# 과거 Coupang 실제 값에서 역산한 고정 단가를 여기 넣어두면 회사가 달라져도
# 그대로 출력되어 "쿠팡 고정" 현상이 발생한다. 이제는 항상 weighted_unit_prices
# (현재 필터링된 line_items 기반 가중 평균)를 사용하도록 비워둔다.
# 특정 SKU에 대해 수동 override가 필요할 때만 키를 추가한다.
# ─────────────────────────────────────────────────────────────────────────────
FORCED_UNIT_PRICES: dict[str, float] = {}

# ─────────────────────────────────────────────────────────────────────────────
# 색상 팔레트
# ─────────────────────────────────────────────────────────────────────────────
C_HEADER  = "E2EFDA"
C_SUB     = "F2F2F2"
C_WHITE   = "FFFFFF"
C_ORANGE  = "FF8F00"
C_DIVIDER = "E0E0E0"
C_TEXT    = "000000"
C_TOTAL   = "C5E0B3"   # toal 행 배경색 (#C5E0B3)

_BLK       = Side(style="thin", color="000000")
_BORDER    = Border(left=_BLK, right=_BLK, top=_BLK, bottom=_BLK)
_NO_BORDER = Border()


# ─────────────────────────────────────────────────────────────────────────────
# 스타일 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color: str = C_TEXT, bold: bool = False, size: int = 8) -> Font:
    return Font(color=color, bold=bold, size=size, name="맑은 고딕")


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _pre_merge_style(ws, r1, c1, r2, c2, fill=None, border=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if fill   is not None: cell.fill   = fill
            if border is not None: cell.border = border


def _set(ws, r, c, value=None, fill=None, font=None,
         alignment=None, border=None, number_format=None):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        return
    if value         is not None: cell.value         = value
    if fill          is not None: cell.fill          = fill
    if font          is not None: cell.font          = font
    if alignment     is not None: cell.alignment     = alignment
    if border        is not None: cell.border        = border
    if number_format is not None: cell.number_format = number_format


def _add_image(ws, img_dir, fname, col_0idx, row_0idx,
               width_cm, height_cm, offset_cm=0.1):
    """
    col_0idx, row_0idx: 0-based 열/행 인덱스 (openpyxl AnchorMarker 기준)
    width_cm / height_cm: 이미지 크기 (cm 단위)
    offset_cm: 상단 여백 (cm 단위)
    """
    if not _HAS_IMAGE:
        return
    path = os.path.join(img_dir, fname)
    if not os.path.exists(path):
        return
    try:
        img    = XLImage(path)
        w_emu  = int(width_cm  * _CM_TO_EMU)
        h_emu  = int(height_cm * _CM_TO_EMU)
        off_emu = int(offset_cm * _CM_TO_EMU)
        marker = AnchorMarker(col=col_0idx, colOff=0,
                              row=row_0idx, rowOff=off_emu)
        size   = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# 공개 함수
# ─────────────────────────────────────────────────────────────────────────────
def write_project_sheet(
    wb: Workbook,
    proj_results: list,
    company_name: str,
    billing_month: str,
    exchange_rate: Decimal,
    _margin_rate: Decimal,
    invoice_date: str | None,
    bank_name: str,
    line_items: list | None = None,   # noqa: ARG001  (호환용 — 더 이상 사용하지 않음)
    invoice_sku_rows: list[dict] | None = None,
    invoice_rate_row: int | None = None,
    per_proj_invoice_meta: dict[str, dict] | None = None,
    currency: str = "USD",
    min_charge_amount: float = 500_000,
    min_charge_currency: str = "KRW",
) -> None:
    """Invoice 시트 내 각 SKU의 header/subtotal 행을 수식으로 참조한다.
    invoice_sku_rows: [{"sku_name","header_row","subtotal_row"}, ...] — Invoice!B/C/I 참조용
                      (account 모드).
    invoice_rate_row: Invoice 시트의 환율 행 번호 (I{rate_row}) — account 모드.
    per_proj_invoice_meta: per_project 모드에서만 사용.
        {proj_name: {"sheet_title", "sku_rows": [...], "rate_row": int|None}} 구조로
        각 프로젝트 블록이 자기 Invoice 시트를 수식 참조할 수 있도록 한다.
    currency: 'USD' (기본) 또는 'KRW'. KRW 모드에선 환율 메타 행 생략, 모든
              금액 ₩ 표기, 프로젝트별 toal($) 행 숨김, 총 합계는 직접 SUM.
    """
    is_krw = (currency == "KRW")

    ws = wb.create_sheet("Project")
    ws.sheet_view.showGridLines = False

    # ── API 목록: Invoice 시트 순서 그대로 (Tax/VAT 제외는 invoice쪽에서 이미 완료) ─
    # 세금/VAT 판별 — 한글은 substring, 영문은 단어 경계 (Elevation 의 'vat' 오탐 방지).
    _TAX_RE = re.compile(r"세금|\btax\b|\bvat\b", re.IGNORECASE)

    def _is_tax_sku(name: str) -> bool:
        return bool(_TAX_RE.search(name or ""))

    api_list: list[str] = []
    sku_row_map: dict[str, dict] = {}
    if invoice_sku_rows:
        # Invoice 시트에 배치된 순서를 그대로 유지 (사용자 지정 순서 존중)
        for s in invoice_sku_rows:
            nm = s["sku_name"]
            if _is_tax_sku(nm):
                continue
            api_list.append(nm)
            sku_row_map[nm] = {
                "header_row":   s["header_row"],
                "subtotal_row": s["subtotal_row"],
            }
    else:
        # invoice_sku_rows 미전달(per_project 모드) → proj_results 기반 fallback.
        # 어느 프로젝트든 **실제 사용량(usage) > 0** 또는 **청구 금액(subtotal_usd) > 0**
        # 인 SKU 는 컬럼으로 노출. free cap 내에서 전액 소진되어 amount=0 이어도
        # usage 가 있으면 "사용 기록" 차원에서 표시한다.
        _total_sub: dict[str, Decimal] = {}
        _has_usage: dict[str, bool] = {}
        for pr in proj_results:
            for sku_name, sd in pr["skus"].items():
                if _is_tax_sku(sku_name):
                    continue
                _sub = (sd or {}).get("subtotal_usd") or Decimal("0")
                if not isinstance(_sub, Decimal):
                    try:
                        _sub = Decimal(str(_sub))
                    except Exception:
                        _sub = Decimal("0")
                _total_sub[sku_name] = _total_sub.get(sku_name, Decimal("0")) + _sub
                try:
                    _u = int((sd or {}).get("usage") or 0)
                except (TypeError, ValueError):
                    _u = 0
                if _u > 0:
                    _has_usage[sku_name] = True
        api_list = [
            nm for nm, s in _total_sub.items()
            if s > 0 or _has_usage.get(nm)
        ]
        order_map = {name: i for i, name in enumerate(PREFERRED_API_ORDER)}
        api_list.sort(key=lambda n: order_map.get(n, len(PREFERRED_API_ORDER)))

    NUM_APIS = len(api_list)

    # ── 열 인덱스 정의 (3컬럼 per API) ──────────────────────────────────────
    LABEL_COL = 1   # A: 빈 여백
    PROJ_COL  = 2   # B: 프로젝트명

    def api_left(k):  return 3 + 3 * k        # 가격 라벨 좌측
    def api_mid(k):   return 3 + 3 * k + 1    # 가격 라벨 우측 (left와 병합)
    def api_right(k): return 3 + 3 * k + 2    # amount 단독 컬럼

    MAX_COL = 2 + 3 * NUM_APIS   # = api_right(NUM_APIS - 1)

    # ── 열 너비 ───────────────────────────────────────────────────────────────
    # (left_w, mid_w, right_w) per API
    _COL_WIDTHS: dict[str, tuple] = {
        "Dynamic Maps":                                      (9, 5, 12),
        "Basic Data":                                        (9, 5, 12),
        "Contact Data":                                      (9, 5, 12),
        "Atmosphere Data":                                   (9, 5, 12),
        "Find Place":                                        (9, 5, 12),
        "Geocoding":                                         (9, 5, 12),
        "Autocomplete - Per Request":                        (9, 5, 12),
        "Autocomplete without Places Details - Per Session": (9, 6, 12),
        "Places Details":                                    (9, 5, 12),
        "Places - Text Search":                              (9, 5, 12),
        "Query Autocomplete - Per Request":                  (9, 5, 12),
        "Directions":                                        (9, 5, 12),
    }
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 2
    ws.column_dimensions[get_column_letter(PROJ_COL)].width  = 16
    for k, name in enumerate(api_list):
        lw, mw, rw = _COL_WIDTHS.get(name, (9, 5, 12))
        ws.column_dimensions[get_column_letter(api_left(k))].width  = lw
        ws.column_dimensions[get_column_letter(api_mid(k))].width   = mw
        ws.column_dimensions[get_column_letter(api_right(k))].width = rw

    # ── 전체 흰색 배경 초기화 — 실제 콘텐츠 범위만 (9 meta + 7×proj + 10 여유) ──
    wf = _fill(C_WHITE)
    _max_bg_row = 9 + len(proj_results) * 7 + 10
    for _row in ws.iter_rows(min_row=1, max_row=_max_bg_row,
                             min_col=1, max_col=MAX_COL + 2):
        for _cell in _row:
            _cell.fill = wf

    def rh(r, h): ws.row_dimensions[r].height = h

    # ── 날짜 계산 ─────────────────────────────────────────────────────────────
    if invoice_date is None:
        invoice_date = _date.today().strftime("%Y-%m-%d")
    year, month = int(billing_month[:4]), int(billing_month[5:7])
    last_day    = _cal.monthrange(year, month)[1]
    term_str    = f"{billing_month}-01 ~ {billing_month}-{last_day:02d}"

    # ══════════════════════════════════════════════════════════════════════════
    # 상단 이미지 헤더 (Row 1~3)
    # ══════════════════════════════════════════════════════════════════════════
    rh(1, 50)
    rh(2, 0.75)   # 로고와 송장 정보 사이 간격 제거
    rh(3, 0.75)
    rh(4, 0.75)

    for r in range(1, 5):
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill   = _fill(C_WHITE)
            ws.cell(row=r, column=c).border = _NO_BORDER

    _img_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")

    # tt1.png: B1 (col_0idx=1, row_0idx=0), 8.0cm × 1.2cm, offset 0.1cm
    _add_image(ws, _img_dir, "tt1.png",
               col_0idx=1, row_0idx=0,
               width_cm=8.0, height_cm=1.2, offset_cm=0.1)

    # tt2.png: F1 (col_0idx=5, row_0idx=0), 2.83cm × 1.29cm, offset 0.1cm
    _add_image(ws, _img_dir, "tt2.png",
               col_0idx=5, row_0idx=0,
               width_cm=2.83, height_cm=1.29, offset_cm=0.1)

    # ══════════════════════════════════════════════════════════════════════════
    # 송장 요약 정보 (Row 5~8, B열 시작)
    # ══════════════════════════════════════════════════════════════════════════
    # 환율 numeric 셀은 별도로 (PROJ_COL+1 = C열) — 수식에서 $C$8 로 참조 가능
    # Row 5~7: (label, 값 텍스트 — 병합)
    # Row 8:   (label=환율, C8=numeric 환율, D8:MAX_COL=bank 주석)
    meta_rows_text = [
        ("Invoice Date",         f": {invoice_date}"),
        ("Billing Account Name", f": {company_name.title()}"),
        ("Term of Use",          f": {term_str}"),
    ]
    for i, (label, val) in enumerate(meta_rows_text):
        r = 5 + i
        rh(r, 17)
        ws.cell(row=r, column=LABEL_COL).fill   = _fill(C_WHITE)
        ws.cell(row=r, column=LABEL_COL).border = _NO_BORDER

        cell = ws.cell(row=r, column=PROJ_COL, value=label)
        cell.fill      = _fill(C_WHITE)
        cell.border    = _NO_BORDER
        cell.font      = Font(color="555555", bold=True, size=9, name="맑은 고딕")
        cell.alignment = _align("left", "center")

        _pre_merge_style(ws, r, PROJ_COL + 1, r, MAX_COL,
                         fill=_fill(C_WHITE), border=_NO_BORDER)
        ws.merge_cells(start_row=r, start_column=PROJ_COL + 1,
                       end_row=r,   end_column=MAX_COL)
        cell = ws.cell(row=r, column=PROJ_COL + 1, value=val)
        cell.fill      = _fill(C_WHITE)
        cell.border    = _NO_BORDER
        cell.font      = Font(color="111111", bold=False, size=9, name="맑은 고딕")
        cell.alignment = _align("left", "center")

    # ── Row 8: 환율 — USD 모드에서만 표기, KRW 모드에서는 생략 ────────────────
    r_rate = 5 + len(meta_rows_text)   # = 8
    RATE_CELL = f"${get_column_letter(PROJ_COL + 1)}${r_rate}"   # "$C$8" (USD 모드만 참조됨)

    if not is_krw:
        rh(r_rate, 17)

        ws.cell(row=r_rate, column=LABEL_COL).fill   = _fill(C_WHITE)
        ws.cell(row=r_rate, column=LABEL_COL).border = _NO_BORDER

        cell = ws.cell(row=r_rate, column=PROJ_COL, value="환율")
        cell.fill      = _fill(C_WHITE)
        cell.border    = _NO_BORDER
        cell.font      = Font(color="555555", bold=True, size=9, name="맑은 고딕")
        cell.alignment = _align("left", "center")

        # C8: Invoice 시트 환율 셀 참조 (수식) + 숫자 포맷
        if invoice_rate_row is not None:
            rate_value = f"=Invoice!I{invoice_rate_row}"
        else:
            rate_value = float(exchange_rate)
        cell = ws.cell(row=r_rate, column=PROJ_COL + 1, value=rate_value)
        cell.fill          = _fill(C_WHITE)
        cell.border        = _NO_BORDER
        cell.font          = Font(color="111111", bold=True, size=9, name="맑은 고딕")
        cell.alignment     = _align("left", "center")
        cell.number_format = '"\u20a9"#,##0.00'

        # D8:MAX_COL — bank / 기준일 주석
        _pre_merge_style(ws, r_rate, PROJ_COL + 2, r_rate, MAX_COL,
                         fill=_fill(C_WHITE), border=_NO_BORDER)
        ws.merge_cells(start_row=r_rate, start_column=PROJ_COL + 2,
                       end_row=r_rate,   end_column=MAX_COL)
        cell = ws.cell(row=r_rate, column=PROJ_COL + 2,
                       value=f"  ({bank_name} {year}.{month:02d}.{last_day:02d} 최종 송금환율 기준)")
        cell.fill      = _fill(C_WHITE)
        cell.border    = _NO_BORDER
        cell.font      = Font(color="111111", bold=False, size=9, name="맑은 고딕")
        cell.alignment = _align("left", "center")
    else:
        rh(r_rate, 4)   # KRW: 환율 행 자리를 얇은 여백으로

    # Row 9: 여백
    rh(9, 8)

    # ══════════════════════════════════════════════════════════════════════════
    # 프로젝트 블록 (6행 per project, Row 10부터 시작)
    # ══════════════════════════════════════════════════════════════════════════
    cur       = 10
    project_usd_cells: list[str] = []   # 각 프로젝트의 toal($) 셀 참조 — grand total 수식용
    project_krw_cells: list[str] = []   # 각 프로젝트의 toal(₩) 셀 참조 (USD 모드) — 합산용

    hdr_fill  = _fill(C_HEADER)
    data_fill = _fill(C_WHITE)
    tot_fill  = _fill(C_TOTAL)   # toal 행 공통 배경 (#C5E0B3)

    # 프로젝트명 번호 순 정렬 (coupang-01 → coupang-02 ...), 번호 없는 프로젝트(Butter 등)는 맨 뒤
    def _proj_sort_key(p):
        name = p["proj_name"]
        m = re.search(r"-(\d+)", name)
        return (0, name) if m else (1, name)

    proj_results = sorted(proj_results, key=_proj_sort_key)

    def _quote_sheet(name: str) -> str:
        # Excel 수식 시트 참조: 시트명에 ' 가 있으면 '' 로 이스케이프 후 홑따옴표로 감쌈.
        return f"'{(name or '').replace(chr(39), chr(39)*2)}'"

    for pr in proj_results:
        r_name   = cur       # 행1: 서비스명 (녹색)
        r_usage  = cur + 1   # 행2: 사용량   (녹색)
        r_labels = cur + 2   # 행3: 라벨     (WHITE)
        r_vals   = cur + 3   # 행4: 값       (WHITE)
        r_usd    = cur + 4   # 행5: toal($)
        r_krw    = cur + 5   # 행6: toal(₩)

        proj_amount_cells: list[str] = []   # 이 프로젝트의 amount 셀 (E, H, K, ...)

        # ── per_project 모드: 이 프로젝트 전용 Invoice 시트 메타 조회 ─────────
        #   있으면 해당 시트의 sku_rows 로 로컬 맵 구성하고 수식 참조용 sheet prefix 준비.
        #   없으면 account 모드 기본값(전역 sku_row_map + "Invoice" 시트명) 사용.
        _pp_meta = (per_proj_invoice_meta or {}).get(pr.get("proj_name"))
        if _pp_meta:
            _local_sku_row_map: dict[str, dict] = {}
            for _s in (_pp_meta.get("sku_rows") or []):
                _nm = _s.get("sku_name")
                if _nm and not _is_tax_sku(_nm):
                    _local_sku_row_map[_nm] = {
                        "header_row":   _s["header_row"],
                        "subtotal_row": _s["subtotal_row"],
                    }
            _sheet_ref = _quote_sheet(_pp_meta.get("sheet_title") or "Invoice")
        else:
            _local_sku_row_map = sku_row_map
            _sheet_ref = "Invoice"

        for r, h in zip(
            [r_name, r_usage, r_labels, r_vals, r_usd, r_krw],
            [22,     18,      14,       18,     14,    18],
        ):
            rh(r, h)

        # ── Col A: 빈 여백 ────────────────────────────────────────────────────
        for r in range(r_name, r_krw + 1):
            ws.cell(row=r, column=LABEL_COL).fill   = _fill(C_WHITE)
            ws.cell(row=r, column=LABEL_COL).border = _NO_BORDER

        # ── Col B: 프로젝트명 (4행 세로 병합, WHITE) ─────────────────────────
        display_name = re.sub(
            r'\s*-\s*project\s*$', '', pr["proj_name"], flags=re.IGNORECASE
        ).strip()

        _pre_merge_style(ws, r_name, PROJ_COL, r_vals, PROJ_COL,
                         fill=data_fill, border=_BORDER)
        try:
            ws.merge_cells(start_row=r_name, start_column=PROJ_COL,
                           end_row=r_vals,   end_column=PROJ_COL)
        except Exception:
            pass
        cell = ws.cell(row=r_name, column=PROJ_COL, value=display_name)
        cell.fill      = data_fill
        cell.font      = _font(C_TEXT, bold=True, size=9)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border    = _BORDER

        # ── API별 4행 격자 (3컬럼 구조) ──────────────────────────────────────
        # 이 프로젝트에서 실제 사용/청구한 API 만 왼쪽부터 순서대로 배치한다.
        # (공통 api_list 그대로 쓰면 미사용 API 자리가 빈 칸으로 남아 레이아웃이
        #  비고 toal 값 열이 허전하게 외떨어져 보인다.)
        used_apis: list[str] = []
        for _nm in api_list:
            _sd_ck = pr["skus"].get(_nm, {})
            try:
                _u_ck = int(_sd_ck.get("usage") or 0)
            except (TypeError, ValueError):
                _u_ck = 0
            _sub_ck = _sd_ck.get("subtotal_usd") or Decimal("0")
            _krw_ck = _sd_ck.get("final_krw")    or Decimal("0")
            if _u_ck > 0 or _sub_ck > 0 or _krw_ck > 0:
                used_apis.append(_nm)

        for k, api_name in enumerate(used_apis):
            sd = pr["skus"].get(
                api_name,
                {"usage": 0, "subtotal_usd": Decimal("0"), "final_krw": Decimal("0")},
            )
            usage   = sd["usage"]

            lc = api_left(k)
            mc = api_mid(k)
            rc = api_right(k)

            lc_letter = get_column_letter(lc)
            rc_letter = get_column_letter(rc)

            # Invoice 참조 행 (수식용) — per_project 모드면 로컬 맵 사용.
            inv_rows = _local_sku_row_map.get(api_name, {})
            inv_header   = inv_rows.get("header_row")
            inv_subtotal = inv_rows.get("subtotal_row")

            # 행1: 서비스명 (C:E 3열 병합, 녹색, 굵게)
            short = API_DISPLAY_NAMES.get(api_name) or (
                api_name.replace(" - ", "\n").replace(" without ", "\nw/o ")
                if len(api_name) > 14 else api_name
            )
            _pre_merge_style(ws, r_name, lc, r_name, rc,
                             fill=hdr_fill, border=_BORDER)
            try:
                ws.merge_cells(start_row=r_name, start_column=lc,
                               end_row=r_name,   end_column=rc)
            except Exception:
                pass
            cell = ws.cell(row=r_name, column=lc, value=short)
            cell.fill      = hdr_fill
            cell.font      = _font(C_TEXT, bold=True, size=8)
            cell.alignment = _align("center", "center", wrap=True)
            cell.border    = _BORDER

            # 행2: 사용량 (C:E 3열 병합, 녹색, 우측 정렬)
            _pre_merge_style(ws, r_usage, lc, r_usage, rc,
                             fill=hdr_fill, border=_BORDER)
            try:
                ws.merge_cells(start_row=r_usage, start_column=lc,
                               end_row=r_usage,   end_column=rc)
            except Exception:
                pass
            cell = ws.cell(row=r_usage, column=lc,
                           value=usage)
            cell.fill          = hdr_fill
            cell.font          = _font(C_TEXT, size=8)
            cell.alignment     = _align("right", "center")
            cell.border        = _BORDER
            cell.number_format = '#,##0;;""'

            # 행3: 라벨 (C:D 병합=monthly unit price($), E=amount, WHITE)
            _pre_merge_style(ws, r_labels, lc, r_labels, mc,
                             fill=data_fill, border=_BORDER)
            try:
                ws.merge_cells(start_row=r_labels, start_column=lc,
                               end_row=r_labels,   end_column=mc)
            except Exception:
                pass
            cell = ws.cell(row=r_labels, column=lc,
                           value="monthly unit price($)")
            cell.fill      = data_fill
            cell.font      = _font(C_TEXT, size=7)
            cell.alignment = _align("center", "center", wrap=True)
            cell.border    = _BORDER

            _set(ws, r_labels, rc,
                 value="amount",
                 fill=data_fill,
                 font=_font(C_TEXT, size=7),
                 alignment=_align("center", "center"),
                 border=_BORDER)

            # 행4: 값 (C:D 병합=단가 수식, E=금액 수식, WHITE)
            _pre_merge_style(ws, r_vals, lc, r_vals, mc,
                             fill=data_fill, border=_BORDER)
            try:
                ws.merge_cells(start_row=r_vals, start_column=lc,
                               end_row=r_vals,   end_column=mc)
            except Exception:
                pass

            # 단가·amount 계산 경로:
            #   (A) Invoice 셀 참조 가능 (account 모드) → 단순 분수 수식 체인 유지.
            #       "단가 = Invoice I(소계) / Invoice C(Usage)". 무료로 전액
            #       소진된 경우(usage>0, subtotal=0)는 결과가 자연스레 0 이 되어
            #       별도 가드 불필요. usage=0 인 edge case(subtotal>0, usage=0
            #       같은 비정상 데이터) 만 IF 로 DIV/0 방지.
            #   (B) per_project 모드 → Invoice 시트가 여러 개라 참조 불가.
            #       subtotal(=waterfall 결과, Decimal) 을 amount 에 **직접** 기록해
            #       float round-trip 없이 정밀도 유지. 단가는 subtotal/usage.
            _amount_value: object
            if inv_header is not None and inv_subtotal is not None:
                try:
                    _u_inv = int(usage or 0)
                except (TypeError, ValueError):
                    _u_inv = 0
                if _u_inv > 0:
                    unit_price_val: object = (
                        f"={_sheet_ref}!$I${inv_subtotal}/{_sheet_ref}!$C${inv_header}"
                    )
                else:
                    # 사용량 0 → DIV/0 방지용 가드
                    unit_price_val = (
                        f"=IF({_sheet_ref}!$C${inv_header}>0,"
                        f"{_sheet_ref}!$I${inv_subtotal}/{_sheet_ref}!$C${inv_header},0)"
                    )
                _amount_value = None   # 아래 수식 경로에서 설정
            else:
                _sub_dec   = sd.get("subtotal_usd", Decimal("0"))
                _final_dec = sd.get("final_krw",    Decimal("0"))
                _u         = int(usage or 0)
                if _u > 0:
                    if is_krw:
                        unit_price_val = float(_final_dec) / _u
                        _amount_value  = float(_final_dec)
                    else:
                        unit_price_val = float(_sub_dec) / _u
                        _amount_value  = float(_sub_dec)
                else:
                    unit_price_val = 0
                    _amount_value  = 0
            # 통화별 단가 포맷: USD는 소수 3자리, KRW는 정수(원 단위)
            _up_fmt = '"\u20a9"#,##0;-"₩"#,##0;"₩"0' if is_krw else '#,##0.000;-#,##0.000;0'
            cell = ws.cell(row=r_vals, column=lc, value=unit_price_val)
            cell.fill          = data_fill
            cell.font          = _font(C_TEXT, size=8)
            cell.alignment     = _align("right", "center")
            cell.border        = _BORDER
            cell.number_format = _up_fmt

            # amount 셀: usage 셀 × 단가 셀 (같은 블록 내 셀 참조)
            usage_ref      = f"{lc_letter}{r_usage}"
            unit_price_ref = f"{lc_letter}{r_vals}"
            amount_ref     = f"{rc_letter}{r_vals}"
            proj_amount_cells.append(amount_ref)
            _amt_fmt = '"\u20a9"#,##0;-"₩"#,##0;"₩"0' if is_krw else '#,##0;-#,##0;0'
            if _amount_value is None:
                _amount_value = f"={usage_ref}*{unit_price_ref}"
            _set(ws, r_vals, rc,
                 value=_amount_value,
                 fill=data_fill,
                 font=_font(C_TEXT, size=8),
                 alignment=_align("right", "center"),
                 border=_BORDER,
                 number_format=_amt_fmt)

        # ── 행5: 프로젝트 total — USD 모드는 toal($), KRW 모드는 toal(₩) ──
        _pre_merge_style(ws, r_usd, PROJ_COL, r_usd, PROJ_COL + 2,
                         fill=tot_fill, border=_BORDER)
        try:
            ws.merge_cells(start_row=r_usd, start_column=PROJ_COL,
                           end_row=r_usd,   end_column=PROJ_COL + 2)
        except Exception:
            pass
        _tot_label = "toal(\u20a9)" if is_krw else "toal($)"
        cell = ws.cell(row=r_usd, column=PROJ_COL, value=_tot_label)
        cell.fill      = tot_fill
        cell.font      = _font(C_TEXT, bold=True, size=8)
        cell.alignment = _align("right", "center")
        cell.border    = _BORDER

        # total = ROUND(SUM(amount 셀 전부), N)
        # — USD 모드: 2 자리 소수 유지 ($316.06 같이 실수 값 보존).
        #   정수(0)로 반올림하면 프로젝트별로 최대 $0.49 오차가 생기고,
        #   그 정수 합에 환율을 곱하면 프로젝트 수만큼 ₩1,500 단위 오차가
        #   청구 대상 금액에 누적된다.
        # — KRW 모드: 1원 단위 정수.
        usd_col_letter = get_column_letter(PROJ_COL + 3)
        usd_cell_ref   = f"{usd_col_letter}{r_usd}"
        project_usd_cells.append(usd_cell_ref)
        if proj_amount_cells:
            _rd = 0 if is_krw else 2
            usd_formula = f"=ROUND(SUM({','.join(proj_amount_cells)}),{_rd})"
        else:
            usd_formula = 0
        _tot_fmt = '"\u20a9"#,##0' if is_krw else '"$"#,##0.00'
        _set(ws, r_usd, PROJ_COL + 3,
             value=usd_formula,
             fill=tot_fill,
             font=_font(C_TEXT, bold=True, size=9),
             alignment=_align("right", "center"),
             border=_BORDER,
             number_format=_tot_fmt)

        # E 이후: 흰색 배경, 테두리 없음
        for c in range(PROJ_COL + 4, MAX_COL + 1):
            _set(ws, r_usd, c, fill=data_fill, border=_NO_BORDER)

        # ── 행6: toal(₩) — USD 모드에서만 렌더 (KRW 모드는 위 한 줄이 이미 ₩) ─
        if not is_krw:
            _pre_merge_style(ws, r_krw, PROJ_COL, r_krw, PROJ_COL + 2,
                             fill=tot_fill, border=_BORDER)
            try:
                ws.merge_cells(start_row=r_krw, start_column=PROJ_COL,
                               end_row=r_krw,   end_column=PROJ_COL + 2)
            except Exception:
                pass
            cell = ws.cell(row=r_krw, column=PROJ_COL, value="toal(\u20a9)")
            cell.fill      = tot_fill
            cell.font      = _font(C_TEXT, bold=True, size=8)
            cell.alignment = _align("right", "center")
            cell.border    = _BORDER

            # toal(₩) = ROUND(toal($) × 환율, 0) — $C$8 참조
            _krw_col_letter = get_column_letter(PROJ_COL + 3)
            project_krw_cells.append(f"{_krw_col_letter}{r_krw}")
            _set(ws, r_krw, PROJ_COL + 3,
                 value=f"=ROUND({usd_cell_ref}*{RATE_CELL},0)",
                 fill=tot_fill,
                 font=_font(C_TEXT, bold=True, size=9),
                 alignment=_align("right", "center"),
                 border=_BORDER,
                 number_format="\u20a9#,##0")

            for c in range(PROJ_COL + 4, MAX_COL + 1):
                _set(ws, r_krw, c, fill=data_fill, border=_NO_BORDER)

            cur = r_krw + 1
        else:
            # KRW 모드: 6번째 행 자리를 여백으로 축소
            rh(r_krw, 4)
            cur = r_krw + 1

        # ── 프로젝트 간 구분선 ────────────────────────────────────────────────
        rh(cur, 5)
        for c in range(1, MAX_COL + 1):
            ws.cell(row=cur, column=c).fill = _fill(C_SUB)
        cur += 1

    # ══════════════════════════════════════════════════════════════════════════
    # 최하단 합계 행 (green #C5E0B3, B:E 범위만)
    # ══════════════════════════════════════════════════════════════════════════
    # 월 최소 사용비용 (계정별 설정). USD 로 설정돼 있으면 환율로 KRW 환산.
    # amount ≤ 0 이면 룰 자체를 적용하지 않는다(면제 계약 계정용).
    try:
        _mc_amt = float(min_charge_amount)
    except (TypeError, ValueError):
        _mc_amt = 0.0
    if _mc_amt <= 0:
        _MIN_CHARGE_KRW = 0
    elif min_charge_currency == "USD":
        _MIN_CHARGE_KRW = int(round(_mc_amt * float(exchange_rate)))
    else:
        _MIN_CHARGE_KRW = int(round(_mc_amt))

    try:
        _proj_usd_total = sum(
            float(pr.get("total_usd") or 0) for pr in proj_results
        )
    except Exception:
        _proj_usd_total = 0.0
    if is_krw:
        _predicted_krw = round(_proj_usd_total)
    else:
        _predicted_krw = round(_proj_usd_total * float(exchange_rate))
    _min_charge_applied = (
        _MIN_CHARGE_KRW > 0 and 0 < _predicted_krw < _MIN_CHARGE_KRW
    )

    rh(cur, 5)
    cur += 1

    if _min_charge_applied:
        r_min_charge = cur
        r_total      = cur + 1
        r_vat_notice = cur + 2
        rh(r_min_charge, 22)
    else:
        r_min_charge = None
        r_total      = cur
        r_vat_notice = cur + 1
    rh(r_total,      22)
    rh(r_vat_notice, 14)

    # A열 및 F열 이후 전체 흰색으로 초기화 (주황색 제거)
    _reset_rows = [r_total, r_vat_notice]
    if r_min_charge is not None:
        _reset_rows.append(r_min_charge)
    for r in _reset_rows:
        for c in range(1, MAX_COL + 2):
            ws.cell(row=r, column=c).fill   = _fill(C_WHITE)
            ws.cell(row=r, column=c).border = _NO_BORDER

    # ── 월최소사용비용(KRW) 행 (조건부) ──
    if r_min_charge is not None:
        _pre_merge_style(ws, r_min_charge, PROJ_COL, r_min_charge, PROJ_COL + 2,
                         fill=tot_fill, border=_BORDER)
        try:
            ws.merge_cells(start_row=r_min_charge, start_column=PROJ_COL,
                           end_row=r_min_charge,   end_column=PROJ_COL + 2)
        except Exception:
            pass
        cell = ws.cell(row=r_min_charge, column=PROJ_COL, value="월최소사용비용(KRW)")
        cell.fill      = tot_fill
        cell.font      = Font(color=C_TEXT, bold=True, size=10, name="맑은 고딕")
        cell.alignment = _align("center", "center")
        cell.border    = _BORDER
        _set(ws, r_min_charge, PROJ_COL + 3,
             value=_MIN_CHARGE_KRW,
             fill=tot_fill,
             font=Font(color=C_TEXT, bold=True, size=10, name="맑은 고딕"),
             alignment=_align("right", "center"),
             border=_BORDER,
             number_format="\u20a9#,##0")

    # B:D 병합 → "청구 대상 금액" 라벨 (#C5E0B3, 검정 굵게)
    _pre_merge_style(ws, r_total, PROJ_COL, r_total, PROJ_COL + 2,
                     fill=tot_fill, border=_BORDER)
    try:
        ws.merge_cells(start_row=r_total, start_column=PROJ_COL,
                       end_row=r_total,   end_column=PROJ_COL + 2)
    except Exception:
        pass
    cell = ws.cell(row=r_total, column=PROJ_COL, value="청구 대상 금액")
    cell.fill      = tot_fill
    cell.font      = Font(color=C_TEXT, bold=True, size=10, name="맑은 고딕")
    cell.alignment = _align("center", "center")
    cell.border    = _BORDER

    # E열 → 청구 대상 금액
    #   - 최소사용비용 적용: 고정 ₩500,000
    #   - 그 외:
    #       USD 모드: SUM(프로젝트별 toal₩)  — 각 프로젝트 KRW 는 이미
    #         ROUND(usd × rate, 0) 이므로 단순 합산. (Invoice 시트 합과
    #         동일 반올림 경로 유지 — per-sheet 반올림 후 합산)
    #       KRW 모드: ROUND(SUM(프로젝트별 toal₩), 0) — 환율 곱 없음
    if _min_charge_applied:
        grand_krw_formula = _MIN_CHARGE_KRW
    elif is_krw and project_usd_cells:
        grand_krw_formula = f"=ROUND(SUM({','.join(project_usd_cells)}),0)"
    elif project_krw_cells:
        grand_krw_formula = f"=SUM({','.join(project_krw_cells)})"
    else:
        grand_krw_formula = 0
    _set(ws, r_total, PROJ_COL + 3,
         value=grand_krw_formula,
         fill=tot_fill,
         font=Font(color="FF0000", bold=True, size=10, name="맑은 고딕"),
         alignment=_align("right", "center"),
         border=_BORDER,
         number_format="\u20a9#,##0")

    # 부가세 안내 — E열 바로 아래 (흰색 배경, 테두리 없음, 우측 정렬, 작은 글씨)
    _set(ws, r_vat_notice, PROJ_COL + 3,
         value="(부가세 별도)",
         fill=_fill(C_WHITE),
         font=Font(color="555555", bold=False, size=7, name="맑은 고딕"),
         alignment=_align("right", "center"),
         border=_NO_BORDER)

    # ── 인쇄 설정 ─────────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
