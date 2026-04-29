import io
from decimal import Decimal
from typing import Any
from billing.models import Sku, SkuTier, UsageRow

_UNLIMITED_CAP = 999_999_999

# 하드코딩 SKU 화이트리스트는 사용하지 않는다.
# 회사마다 Google Maps Platform 에서 사용하는 제품(SKU) 이 모두 다르므로,
# "정답지 12개" 같은 고정 리스트로 필터링하면 Ground K 에만 있는
# Directions Advanced / Time Zone 이나, 다른 계정의 Distance Matrix /
# Places API(New) 계열 등이 자동으로 빠져 버린다. 노출 여부의 기준은
# **선택된 계정의 CSV 에 실제 사용량(usage>0) 이 찍혀 있는지** 하나로 통일.


def filter_canonical_line_items(line_items, canonical=None):
    """usage > 0 인 line_items 만 반환 (이름은 하위 호환용).

    과거엔 canonical 화이트리스트 대조 + usage>0 2중 필터였으나,
    이제는 CSV 실측값 하나를 기준으로 한다. `canonical` 파라미터는
    과거 호출부 호환용으로 남겨 두었을 뿐 동작에 영향이 없다.
    """
    del canonical  # 하위 호환 유지용 — 동작에 사용하지 않음
    return [
        it for it in line_items
        if int(getattr(it, "total_usage", 0) or 0) > 0
    ]


def filter_canonical_proj_results(proj_results, canonical=None):
    """proj_results 의 각 proj.skus 중 usage>0 만 남기고, 빈 proj 제거.

    회사별 사용 SKU 차이를 그대로 반영 — 하드코딩 화이트리스트 없음.
    """
    del canonical
    cleaned = []
    for pr in proj_results:
        new_skus = {
            nm: v for nm, v in pr.get("skus", {}).items()
            if int(v.get("usage", 0) or 0) > 0
        }
        total_usd = sum((v.get("subtotal_usd") or 0) for v in new_skus.values())
        total_krw = sum((v.get("final_krw")    or 0) for v in new_skus.values())
        cleaned.append({
            "proj_id":   pr.get("proj_id"),
            "proj_name": pr.get("proj_name"),
            "skus":      new_skus,
            "total_usd": total_usd,
            "total_krw": total_krw,
        })
    # 해당 기간에 쓴 SKU 가 하나도 없는 프로젝트는 제거
    return [pr for pr in cleaned if pr["skus"]]

def build_sku_master_from_usage(usage_rows, price_list_file) -> dict[str, Sku]:
    """업로드된 CSV 의 usage_rows 에서 (sku_id, sku_name) 을 추출하고,
    Price List(xlsx) 에서 각 sku_name 의 free_cap / tier 단가를 매칭하여
    sku_master 딕셔너리를 **동적으로** 구성한다.

    단일 진실 소스 원칙:
      - "어떤 SKU 가 사용됐는가" → 매달 업로드되는 CSV 고지서
      - "각 SKU 의 단가/무료한도"   → Price List 엑셀

    master_data.csv 같은 중간 정의 파일에 의존하지 않는다. 따라서 신규
    SKU 가 CSV 에 등장해도 Price List 에만 올라가 있으면 자동 집계된다.

    - price_list_file 이 None / 파싱 실패 시 빈 dict 반환 (엔진이 아무것도
      처리하지 않아 명확한 실패 신호가 됨).
    - Price List 에 매칭되지 않는 SKU 는 skip. 이런 SKU 는 detect_missing_skus()
      로 별도 탐지해 UI 경고를 띄운다.
    """
    if price_list_file is None:
        return {}
    try:
        pl_tiers = get_sku_tiers_from_price_list(price_list_file)
    except Exception:
        return {}
    if not pl_tiers:
        return {}

    sid_to_name: dict[str, str] = {}
    for r in usage_rows:
        sid = getattr(r, "sku_id", "") or ""
        nm = (getattr(r, "sku_name", "") or "").strip()
        if sid and nm and sid not in sid_to_name:
            sid_to_name[sid] = nm

    sku_master: dict[str, Sku] = {}
    for sid, nm in sid_to_name.items():
        info = pl_tiers.get(nm)
        if not info:
            continue
        tiers_obj = [
            SkuTier(tier_number=tn, tier_limit=lim, tier_cpm=cpm)
            for (tn, lim, cpm) in info["tiers"]
        ]
        sku_master[sid] = Sku(
            sku_id=sid, sku_name=nm,
            is_billable=True, category="",
            free_usage_cap=int(info["free_cap"] or 0),
            tiers=tiers_obj,
        )
    return sku_master


def detect_missing_skus(usage_rows, sku_master) -> list[tuple[str, str]]:
    """usage_rows 중 sku_master 에 키가 없는 (sku_id, sku_name) 튜플 목록.

    매달 업로드되는 CSV 가 정의 소스이므로, CSV 에 존재하지만 sku_master
    (master_data.csv + Price List fallback) 어디에서도 매칭되지 않은 SKU 는
    엔진이 처리하지 못한 채 조용히 버려지게 된다. 이를 드러내 UI 경고를
    띄울 수 있도록 탐지만 해서 반환한다.

    반환 형식: sorted list of (sku_id, sku_name) — name 기준 안정 정렬.
    """
    seen: dict[tuple[str, str], bool] = {}
    for r in usage_rows:
        sid = getattr(r, "sku_id", "") or ""
        if not sid or sid in sku_master:
            continue
        nm = (getattr(r, "sku_name", "") or "").strip()
        seen[(sid, nm)] = True
    return sorted(seen.keys(), key=lambda x: (x[1].lower(), x[0]))


def load_sku_master(rows: list[dict[str, Any]]) -> dict[str, Sku]:
    sku_map = {}
    for row in rows:
        sid = row["sku_id"]
        if sid not in sku_map:
            sku_map[sid] = Sku(sku_id=sid, sku_name=row["sku_name"], is_billable=row["is_billable"], 
                               category=row["category"], free_usage_cap=row["free_usage_cap"], tiers=[])
        if row.get("tier_number") is not None:
            sku_map[sid].tiers.append(SkuTier(tier_number=int(row["tier_number"]), 
                                             tier_limit=row["tier_limit"], 
                                             tier_cpm=Decimal(str(row["tier_cpm"]))))
    for s in sku_map.values(): s.tiers.sort(key=lambda t: t.tier_number)
    return sku_map

def detect_price_list_currency(price_list_file) -> str:
    """GMP Price List 의 tier 단가 값을 보고 통화를 추정.

    판정 기준: tier 단가 컬럼(D~H, rows 4~50) 의 숫자값을 수집해
    최댓값이 100 이상이면 'KRW', 그 외는 'USD'.
      - USD 단가 예시: 0.15 ~ 30  (최대 30 내외)
      - KRW 단가 예시: 800 ~ 50000
    두 분포의 오버랩이 없어 단순 임계치로 안전하게 구분 가능.

    인자: 파일 경로(str | Path) 또는 file-like 객체 또는 bytes.
    반환: 'USD' (기본) 또는 'KRW'.
    """
    import io
    from openpyxl import load_workbook
    try:
        if hasattr(price_list_file, "read"):
            price_list_file.seek(0)
            wb = load_workbook(io.BytesIO(price_list_file.read()), data_only=True)
        elif isinstance(price_list_file, (bytes, bytearray)):
            wb = load_workbook(io.BytesIO(price_list_file), data_only=True)
        else:
            wb = load_workbook(str(price_list_file), data_only=True)
    except Exception:
        return "USD"

    ws = wb.worksheets[0]
    prices: list[float] = []
    for row in ws.iter_rows(min_row=4, max_row=50,
                            min_col=4, max_col=9, values_only=True):
        for v in row:
            if isinstance(v, (int, float)) and v > 0:
                prices.append(float(v))

    if not prices:
        return "USD"
    return "KRW" if max(prices) >= 100 else "USD"


def get_free_caps_from_price_list(price_list_file) -> dict[str, int]:
    """GMP Price List 엑셀에서 `SKU 이름 → 무료 제공량(int)` 맵.

    판정:
      - Col A: SKU name
      - Col C: Free Usage Cap (정수, 'Unlimited', 또는 헤더 텍스트)
    'Unlimited' 은 매우 큰 수 (_UNLIMITED_CAP) 로 치환.
    숫자 아닌 값은 제외.

    Price List 는 master_data.csv 에 없는 SKU(예: Static Maps, Places API
    Nearby Search Enterprise 등) 도 포함하므로, per_project 모드에서 계정
    무료 제공량 owner 판정 시 master 대신 이 맵을 참조하면 누락 없이 처리
    가능하다.
    """
    import io
    from openpyxl import load_workbook
    try:
        if hasattr(price_list_file, "read"):
            price_list_file.seek(0)
            wb = load_workbook(io.BytesIO(price_list_file.read()), data_only=True)
        elif isinstance(price_list_file, (bytes, bytearray)):
            wb = load_workbook(io.BytesIO(price_list_file), data_only=True)
        else:
            wb = load_workbook(str(price_list_file), data_only=True)
    except Exception:
        return {}

    ws = wb.worksheets[0]
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=200,
                            min_col=1, max_col=3, values_only=True):
        name = row[0]
        cap  = row[2]
        if not isinstance(name, str):
            continue
        name_clean = name.strip()
        if not name_clean:
            continue
        if isinstance(cap, (int, float)) and cap > 0:
            out[name_clean] = int(cap)
        elif isinstance(cap, str) and "unlimited" in cap.lower():
            out[name_clean] = _UNLIMITED_CAP
    return out


def get_sku_tiers_from_price_list(price_list_file) -> dict[str, dict]:
    """GMP Price List 에서 SKU 별 tier 정보 추출.

    반환: {
      "SKU 이름(strip)": {
        "free_cap": int,   # Col C 값 (Unlimited → _UNLIMITED_CAP, 숫자 외 → 0)
        "tiers": [(tier_number, tier_limit, tier_cpm_Decimal), ...],
      },
      ...
    }

    Tier 한도는 GMP 표준 체계(100K / 500K / 1M / 5M / None) 를 Col D..H 에
    일대일 매핑. 가격 셀이 숫자 아닌 경우("별도문의" 등) 해당 tier 는 제외.

    master_data.csv 에 없는 SKU 를 Price List 기반으로 보완해 Python 측
    waterfall 결과가 Excel SUMIF 결과와 일치하도록 하는 용도.
    """
    import io
    from openpyxl import load_workbook
    try:
        if hasattr(price_list_file, "read"):
            price_list_file.seek(0)
            wb = load_workbook(io.BytesIO(price_list_file.read()), data_only=True)
        elif isinstance(price_list_file, (bytes, bytearray)):
            wb = load_workbook(io.BytesIO(price_list_file), data_only=True)
        else:
            wb = load_workbook(str(price_list_file), data_only=True)
    except Exception:
        return {}

    # 표준 tier 한도: (1,100K) (2,500K) (3,1M) (4,5M) (5,∞) 순서로 D,E,F,G,H 매핑
    _STD_LIMITS = [100_000, 500_000, 1_000_000, 5_000_000, None]

    ws = wb.worksheets[0]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=8,
                            values_only=True):
        name = row[0]
        if not isinstance(name, str):
            continue
        name_clean = name.strip()
        if not name_clean:
            continue
        cap_raw = row[2]
        if isinstance(cap_raw, (int, float)) and cap_raw > 0:
            free_cap = int(cap_raw)
        elif isinstance(cap_raw, str) and "unlimited" in cap_raw.lower():
            free_cap = _UNLIMITED_CAP
        else:
            free_cap = 0

        tiers = []
        # Col D..H → tier 1..5
        for i, price_cell in enumerate(row[3:8]):
            if isinstance(price_cell, (int, float)):
                tiers.append((
                    i + 1,
                    _STD_LIMITS[i],
                    Decimal(str(float(price_cell))),
                ))
            # 숫자 아닌 셀은 skip (빈 셀, "별도문의" 등)
        if not tiers:
            continue

        # 중복 이름(가끔 Price List 내 동일 명 재등장) 은 첫 행 우선 유지.
        if name_clean not in out:
            out[name_clean] = {"free_cap": free_cap, "tiers": tiers}
    return out


def get_billable_sku_names(price_list_file) -> set[str]:
    """GMP Price List 엑셀에서 tier 단가가 하나라도 > 0 인 SKU명 집합을 반환.

    판정: rows 4-50, col A = SKU name, cols D-H = tier prices.
    하나라도 양수 단가가 있으면 billable SKU로 간주.
    파일을 읽을 수 없으면 빈 set 반환 (= 필터링 없음).
    """
    import io
    from openpyxl import load_workbook
    try:
        if hasattr(price_list_file, "read"):
            price_list_file.seek(0)
            wb = load_workbook(io.BytesIO(price_list_file.read()), data_only=True)
        elif isinstance(price_list_file, (bytes, bytearray)):
            wb = load_workbook(io.BytesIO(price_list_file), data_only=True)
        else:
            wb = load_workbook(str(price_list_file), data_only=True)
    except Exception:
        return set()

    ws = wb.worksheets[0]
    billable: set[str] = set()
    for row in ws.iter_rows(min_row=4, max_row=50, min_col=1, max_col=9, values_only=False):
        sku_name_cell = row[0]  # col A
        sku_name = sku_name_cell.value
        if not sku_name or not isinstance(sku_name, str):
            continue
        sku_name = sku_name.strip()
        if not sku_name:
            continue
        # cols D-H = indices 3-7 (0-based within the row slice starting at col 1)
        tier_cells = row[3:8]  # cols D(4) through H(8), 0-based index 3..7
        for cell in tier_cells:
            v = cell.value
            if isinstance(v, (int, float)) and v > 0:
                billable.add(sku_name)
                break
    return billable


def load_usage_rows(rows: list[dict[str, Any]]) -> list[UsageRow]:
    from decimal import Decimal as _D
    result = []
    for r in rows:
        krw_val = r.get("cost_krw")
        cost_krw = _D(str(round(float(krw_val), 4))) if krw_val is not None else None
        result.append(UsageRow(
            billing_month=r["billing_month"],
            project_id=r["project_id"],
            project_name=r.get("project_name", r["project_id"]),
            sku_id=r["sku_id"],
            sku_name=r.get("sku_name", "") or "",
            usage_amount=int(r["usage_amount"]),
            cost_krw=cost_krw,
            unit_price=r.get("unit_price"),
        ))
    return result

def load_exchange_rate(row: dict[str, Any]) -> Decimal:
    return Decimal(str(row["usd_to_krw"]))