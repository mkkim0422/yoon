"""
generate_invoice_excel.py

요구사항:
  - billing.csv (UTF-8) 에서 iloc 로 col 2(프로젝트), col 7(SKU 설명), col 14(사용량) 접근
  - 데이터는 skiprows=9 (index 8 = 헤더, index 9 = 첫 데이터 행)
  - SKU 설명 Exact Match 로 40개 API 에 대해서만 집계
  - 출력: 1 개의 xlsx 파일 (Invoice + Project 두 시트)

사용:
  py generate_invoice_excel.py                # 전체 합산 파일 생성
  py generate_invoice_excel.py --project NAME # 특정 프로젝트만
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

API_LIST = [
    "Dynamic Maps", "Directions", "Geocoding",
    "Autocomplete - Per Request",
    "Autocomplete without Places Details - Per Session",
    "Query Autocomplete - Per Request",
    "Places Details", "Basic Data", "Contact Data", "Atmosphere Data",
    "Places - Text Search", "Find Place",
    "Static Maps", "Static Street View", "Street View Metadata", "Aerial View",
    "Dynamic Street View", "Elevation",
    "Places - Nearby Search", "Find Current Place", "Places Photo",
    "Distance Matrix", "Directions Advanced", "Distance Matrix Advanced",
    "Roads - Nearest Road", "Roads - Route Traveled", "Roads - Speed Limits",
    "Address Validation Pro",
    "Solar API", "Air Quality API", "Weather API", "Pollen API",
    "Routes: Compute Route Matrix Pro",
    "RouteOptimization - SingleVehicleRouting",
    "Environment", "Map Tiles",
    "Places API - Legacy", "Basic Data - Legacy",
    "Contact Data - Legacy", "Atmosphere Data - Legacy",
]

SKIP_ROWS = 9
COL_PROJECT = 2
COL_SKU = 7
COL_USAGE = 14


def load_billing(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path, header=None, encoding="utf-8",
                     skiprows=SKIP_ROWS, low_memory=False)
    df = df.iloc[:, [COL_PROJECT, COL_SKU, COL_USAGE]].copy()
    df.columns = ["project", "sku", "usage"]
    df["usage"] = (
        df["usage"].astype(str)
          .str.replace(",", "", regex=False)
          .str.strip()
    )
    df["usage"] = pd.to_numeric(df["usage"], errors="coerce").fillna(0.0)
    df["project"] = df["project"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df = df[df["sku"].isin(API_LIST)]
    return df


def write_invoice_sheet(ws, totals: dict[str, float]) -> None:
    ws.title = "Invoice"
    ws["B8"] = "API 명"
    ws["C8"] = "사용량 합계"
    for c in ("B8", "C8"):
        ws[c].font = Font(bold=True)
        ws[c].fill = PatternFill("solid", fgColor="DDDDDD")
        ws[c].alignment = Alignment(horizontal="center")

    for i, api in enumerate(API_LIST):
        row = 10 + i * 6
        ws.cell(row=row, column=2, value=api)
        ws.cell(row=row, column=3, value=float(totals.get(api, 0)))
        ws.cell(row=row, column=3).number_format = "#,##0"

    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 20


def write_project_sheet(ws, per_project: dict[str, dict[str, float]]) -> None:
    ws.title = "Project"
    ws["A10"] = "프로젝트"
    ws["A10"].font = Font(bold=True)
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A10"].fill = PatternFill("solid", fgColor="DDDDDD")

    for i, api in enumerate(API_LIST):
        start_col = 2 + i * 3
        end_col = start_col + 2
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f"{start_letter}10:{end_letter}10")
        cell = ws.cell(row=10, column=start_col, value=api)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        for col in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    ws.column_dimensions["A"].width = 36

    for r, proj in enumerate(sorted(per_project.keys()), start=11):
        ws.cell(row=r, column=1, value=proj)
        for i, api in enumerate(API_LIST):
            col = 2 + i * 3
            val = per_project[proj].get(api, 0)
            if val:
                c = ws.cell(row=r, column=col, value=float(val))
                c.number_format = "#,##0"


def build_workbook(df: pd.DataFrame, out_path: Path) -> None:
    totals = df.groupby("sku")["usage"].sum().to_dict()
    per_proj_df = (
        df.groupby(["project", "sku"])["usage"].sum().reset_index()
    )
    per_project: dict[str, dict[str, float]] = {}
    for _, r in per_proj_df.iterrows():
        per_project.setdefault(r["project"], {})[r["sku"]] = r["usage"]

    wb = Workbook()
    write_invoice_sheet(wb.active, totals)
    ws2 = wb.create_sheet("Project")
    write_project_sheet(ws2, per_project)
    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default="billing.csv")
    ap.add_argument("--project", default=None,
                    help="특정 프로젝트만 집계 (생략 시 전체)")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    csv_path = Path(args.csv)
    df = load_billing(csv_path)

    if args.project:
        df = df[df["project"] == args.project]
        out = Path(args.out or f"invoice_{args.project}.xlsx")
    else:
        out = Path(args.out or "invoice_total.xlsx")

    build_workbook(df, out)
    print(f"  저장 완료: {out.resolve()}")
    print(f"  집계 행 수: {len(df)}")
    print(f"  매칭된 프로젝트 수: {df['project'].nunique()}")
    print(f"  매칭된 SKU 수: {df['sku'].nunique()} / {len(API_LIST)}")


if __name__ == "__main__":
    main()
