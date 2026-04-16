"""
fill_invoice_from_template.py

template.xlsx 의 디자인·이미지·서식을 그대로 유지한 채
billing.csv 데이터를 result.xlsx 로 주입한다.

  ⚠ pandas.to_excel 사용 금지 — openpyxl 로만 쓴다.
  ⚠ load_workbook(template) 으로 읽어와야 그림/셀 서식이 보존된다.

사용:
  py fill_invoice_from_template.py                      # 프롬프트로 프로젝트 입력
  py fill_invoice_from_template.py --project "LSMNM"    # 인자로 지정
  py fill_invoice_from_template.py --project TOTAL      # 전체 합산
"""
from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ── 설정 ──────────────────────────────────────────────────────────────
TEMPLATE      = "template.xlsx"
CSV_DEFAULT   = "billing.csv"

SKIP_ROWS   = 8    # 스펙: 헤더 9번째 행(index 8) → skiprows=8
COL_PROJECT = 2    # iloc col: 프로젝트 이름
COL_SKU     = 7    # iloc col: SKU 설명
COL_USAGE   = 14   # iloc col: 사용량

API_LIST = [
    "Dynamic Maps", "Directions", "Geocoding",
    "Autocomplete - Per Request",
    "Autocomplete without Places Details - Per Session",
    "Query Autocomplete - Per Request",
    "Places Details", "Basic Data", "Contact Data", "Atmosphere Data",
    "Places - Text Search", "Find Place",
    "Static Maps", "Static Street View", "Street View Metadata", "Aerial View",
    "Dynamic Street View", "Elevation",
    "Places - Nearby Search", "Find Current Place", "Places Photo",
    "Distance Matrix", "Directions Advanced", "Distance Matrix Advanced",
    "Roads - Nearest Road", "Roads - Route Traveled", "Roads - Speed Limits",
    "Address Validation Pro",
    "Solar API", "Air Quality API", "Weather API", "Pollen API",
    "Routes: Compute Route Matrix Pro",
    "RouteOptimization - SingleVehicleRouting",
    "Environment", "Map Tiles",
    "Places API - Legacy", "Basic Data - Legacy",
    "Contact Data - Legacy", "Atmosphere Data - Legacy",
]


# ── 데이터 로드 ────────────────────────────────────────────────────────
def load_usage(csv_path: Path, project: str | None) -> dict[str, float]:
    """billing.csv → {sku: total_usage}. project 지정 시 해당 프로젝트만, 아니면 전체."""
    df = pd.read_csv(csv_path, skiprows=SKIP_ROWS,
                     encoding="utf-8", low_memory=False)

    proj = df.iloc[:, COL_PROJECT].astype(str).str.strip()
    sku  = df.iloc[:, COL_SKU].astype(str).str.strip()
    raw  = (df.iloc[:, COL_USAGE].astype(str)
              .str.replace(",", "", regex=False).str.strip())
    usage = pd.to_numeric(raw, errors="coerce").fillna(0.0)

    mask = sku.isin(API_LIST)
    if project and project.upper() != "TOTAL":
        mask &= (proj == project)

    tmp = pd.DataFrame({"sku": sku[mask], "usage": usage[mask]})
    return tmp.groupby("sku")["usage"].sum().to_dict()


# ── 셀 쓰기 유틸 (병합 안전) ──────────────────────────────────────────
def safe_write(ws, row: int, column: int, value) -> None:
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if (mr.min_row <= row <= mr.max_row
                    and mr.min_col <= column <= mr.max_col):
                ws.cell(row=mr.min_row, column=mr.min_col).value = value
                return
        raise RuntimeError(f"병합 범위 탐색 실패: ({row},{column})")
    cell.value = value


# ── Invoice 시트 주입 ─────────────────────────────────────────────────
def fill_invoice(ws, totals: dict[str, float]) -> int:
    """B10, B16, B22 … (stride 6) 셀의 API 이름을 읽어 C 열에 usage 주입."""
    written = 0
    for i in range(40):
        row = 10 + i * 6
        api = ws.cell(row=row, column=2).value
        if not isinstance(api, str):
            continue
        api = api.strip()
        if api not in API_LIST:
            continue
        safe_write(ws, row=row, column=3, value=float(totals.get(api, 0)))
        written += 1
    return written


# ── Project 시트 주입 ─────────────────────────────────────────────────
_FORMULA_RE = re.compile(r"=Invoice!\$B\$(\d+)", re.IGNORECASE)

def fill_project(ws, invoice_ws, totals: dict[str, float],
                 project_name: str) -> int:
    """Row 10 의 '=Invoice!$B$XX' 수식을 파싱해 각 열이 어떤 API 인지 식별 →
    row 11 에 usage 주입. B10 (병합 B10:B13) 에는 프로젝트명 주입."""
    safe_write(ws, row=10, column=2, value=project_name)

    written = 0
    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=10, column=col).value
        if not (isinstance(v, str) and v.startswith("=")):
            continue
        m = _FORMULA_RE.match(v.strip())
        if not m:
            continue
        b_row = int(m.group(1))
        api = invoice_ws.cell(row=b_row, column=2).value
        if not isinstance(api, str) or api not in API_LIST:
            continue
        safe_write(ws, row=11, column=col, value=float(totals.get(api, 0)))
        written += 1
    return written


# ── 엔트리 ────────────────────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=CSV_DEFAULT)
    ap.add_argument("--template", default=TEMPLATE)
    ap.add_argument("--out", default=None,
                    help="출력 파일명 (기본: result_<project>.xlsx)")
    ap.add_argument("--project", default=None,
                    help="프로젝트명. 'TOTAL' 또는 생략 시 전체 합산.")
    args = ap.parse_args()

    project = args.project
    if project is None:
        project = input("프로젝트명 입력 (전체는 'TOTAL' 또는 엔터): ").strip()
    if not project:
        project = "TOTAL"

    safe_name = re.sub(r"[^\w\-_.]", "_", project)
    out = Path(args.out or f"result_{safe_name}.xlsx")

    # 1) 템플릿 복사 (그림/서식 보존)
    shutil.copy(args.template, out)

    # 2) 집계
    totals = load_usage(Path(args.csv), project)

    # 3) 결과 파일 열어 주입
    wb = load_workbook(out)
    if "Invoice" not in wb.sheetnames or "Project" not in wb.sheetnames:
        raise RuntimeError("템플릿에 'Invoice' / 'Project' 시트가 필요합니다")

    inv_w = fill_invoice(wb["Invoice"], totals)
    prj_w = fill_project(wb["Project"], wb["Invoice"], totals, project)
    wb.save(out)

    print(f"\n  저장 완료: {out.resolve()}")
    print(f"  프로젝트 : {project}")
    print(f"  Invoice 주입 셀: {inv_w}개 / Project 주입 셀: {prj_w}개")
    print(f"  매칭된 SKU : {len(totals)}")
    for sku, v in sorted(totals.items(), key=lambda x: -x[1])[:8]:
        print(f"    {sku:50s} {v:>15,.0f}")


if __name__ == "__main__":
    main()
