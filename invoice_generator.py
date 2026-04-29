"""
invoice_generator.py — Invoice 시트 생성기

레이아웃:
  Col A   : 빈 여백(margin)
  Col B-I : 실제 인보이스 내용 (8개 컬럼)

  Row 1  : 이미지 영역 (tt1 @ B1, tt2 @ I1)
  Row 2  : 빈 스페이서
  Row 3  : Invoice Date
  Row 4  : Billing Account Name
  Row 5  : Term of Use
  Row 6  : 테이블 헤더
  Row 7+ : 데이터 행
  ...    : 합계 / 환율 / 청구금액(KRW)
  ...    : (부가세 별도)
  ...    : tt3 이미지

사용법:
    python invoice_generator.py
    python invoice_generator.py --company coupang --billing-month 2026-03
"""
from __future__ import annotations

import argparse
import calendar
import io
from copy import copy
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from openpyxl.drawing.image import Image as XLImage
    _HAS_IMAGE = True
except Exception:
    _HAS_IMAGE = False

from billing.engine import calculate_billing
from billing.loader import (
    filter_canonical_line_items,
    filter_canonical_proj_results,
    load_sku_master,
    load_usage_rows,
)
from billing.preprocessor import preprocess_usage_file
from project_sheet import write_project_sheet

BASE_DIR   = Path(__file__).parent
ASSETS_DIR = BASE_DIR / "assets"
MASTER_CSV = BASE_DIR / "billing" / "master_data.csv"

from main import SKU_MASTER_ROWS


# ─────────────────────────────────────────────────────────────────────────────
# 레이아웃 상수
# ─────────────────────────────────────────────────────────────────────────────
C1 = 2   # 첫 번째 컨텐츠 열 (B)
CN = 9   # 마지막 컨텐츠 열 (I)


# ─────────────────────────────────────────────────────────────────────────────
# 스타일 상수
# ─────────────────────────────────────────────────────────────────────────────
_THIN       = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="맑은 고딕") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

C_DARK  = "595959"
C_SUB   = "F2F2F2"
C_WHITE = "FFFFFF"


# ─────────────────────────────────────────────────────────────────────────────
# 메인 공개 함수
# ─────────────────────────────────────────────────────────────────────────────
def generate_formatted_invoice(
    line_items: list,
    company_name: str,
    billing_month: str,
    exchange_rate: Decimal,
    margin_rate: Decimal,
    invoice_date: str | None = None,
    output_path: str | Path | None = None,
    bank_name: str = "하나은행",
    proj_results: list | None = None,
    price_list_file=None,
    sku_order: list[str] | None = None,
    currency: str = "USD",
    billable_skus: set[str] | None = None,
    strict_canonical: bool = True,
    billing_mode: str = "account",
    per_project_invoices: list[dict] | None = None,
    min_charge_amount: float = 500_000,
    min_charge_currency: str = "KRW",
    rate_date_str: str | None = None,
    rate_phrase: str = "최종 송금환율 기준",
    rate_extra: str = "",
    include_project_sheet: bool = True,
    subtotal_round: int | None = None,
) -> bytes | None:
    """
    Invoice 시트를 생성한다.

    billing_mode:
      "account"     — 단일 Invoice 시트 (회사 통합 waterfall).
      "per_project" — per_project_invoices 의 항목마다 독립 Invoice 시트 생성.
                      각 시트는 해당 프로젝트 usage 만으로 waterfall 계산된
                      line_items 를 사용한다.
    per_project_invoices: [{"proj_name": str, "line_items": list}, ...]
      billing_mode="per_project" 일 때만 사용. 지정된 순서대로 시트 생성.

    proj_results 전달 시 Project 시트(요약) 도 추가.
    price_list_file 전달 시 GMP Price List 시트도 추가.
    output_path 지정 시 파일 저장 후 None 반환, 아니면 bytes 반환.

    strict_canonical=True (기본):
      total_usage > 0 인 항목만 유지 (하드코딩 화이트리스트 없음).
    """
    def _filter(items):
        if strict_canonical:
            return filter_canonical_line_items(items)
        if billable_skus:
            return [it for it in items if it.sku_name in billable_skus]
        return items

    line_items = _filter(line_items)
    if strict_canonical and proj_results:
        proj_results = filter_canonical_proj_results(proj_results)

    wb = Workbook()

    # ── per_project 모드: 프로젝트별 Invoice 시트 ────────────────────────────
    is_per_project = (
        billing_mode == "per_project"
        and per_project_invoices
    )
    first_sku_rows: list[dict] = []
    first_rate_row: int | None = None
    first_line_items: list = line_items
    # per_project 모드에서 Project 시트가 각 프로젝트 블록의 단가/amount 셀을
    # 해당 프로젝트의 Invoice 시트 셀로 수식 참조할 수 있도록 메타 수집.
    per_proj_invoice_meta: dict[str, dict] = {}

    if is_per_project:
        # 기본 시트를 제거하고 프로젝트별로 새 시트 생성
        _default = wb.active
        wb.remove(_default)
        _n_proj = len(per_project_invoices)
        _prev_sum_refs: list[str] = []   # 이전 시트들의 합계(KRW) 셀 참조 누적
        for idx, entry in enumerate(per_project_invoices):
            proj_name  = str(entry.get("proj_name") or f"Project {idx+1}")
            proj_items = _filter(entry.get("line_items") or [])
            _sheet_title = _safe_sheet_title(proj_name, used=wb.sheetnames)
            ws = wb.create_sheet(_sheet_title)
            _is_last = (idx == _n_proj - 1)
            _ctx = {
                "is_last":          _is_last,
                "prev_sum_refs":    list(_prev_sum_refs),
                "self_sheet_title": _sheet_title,
            }
            # per_project: 각 프로젝트에 배정된 free cap 이 rollover 로 프로젝트
            # 마다 다를 수 있다 (0, 부분값, full cap). Price List SUMIF 는 항상
            # full cap 을 반환하므로, 배정량을 Free Usage 셀에 직접 기록한다.
            _sku_rows, _rate_row, _sum_row = _write_invoice_sheet(
                ws, proj_items, company_name, billing_month,
                invoice_date, exchange_rate, margin_rate, bank_name,
                sku_order=sku_order, currency=currency,
                project_name=proj_name,
                min_charge_amount=min_charge_amount,
                min_charge_currency=min_charge_currency,
                use_item_free_cap=True,
                rate_date_str=rate_date_str,
                rate_phrase=rate_phrase,
                rate_extra=rate_extra,
                per_proj_ctx=_ctx,
                subtotal_round=subtotal_round,
            )
            per_proj_invoice_meta[proj_name] = {
                "sheet_title": _sheet_title,
                "sku_rows":    _sku_rows,
                "rate_row":    _rate_row,
                "sum_row":     _sum_row,
            }
            if _sum_row is not None:
                _safe = _sheet_title.replace("'", "''")
                _prev_sum_refs.append(f"'{_safe}'!I{_sum_row}")
            if idx == 0:
                first_sku_rows   = _sku_rows
                first_rate_row   = _rate_row
                first_line_items = proj_items
    else:
        ws = wb.active
        ws.title = "Invoice"
        # 단일 프로젝트 계정일 때만 Project Name 헤더에 표시
        _proj_name_for_header = ""
        if proj_results and len(proj_results) == 1:
            _proj_name_for_header = str(proj_results[0].get("proj_name") or "")
        first_sku_rows, first_rate_row, _ = _write_invoice_sheet(
            ws, line_items, company_name, billing_month,
            invoice_date, exchange_rate, margin_rate, bank_name,
            sku_order=sku_order, currency=currency,
            project_name=_proj_name_for_header,
            min_charge_amount=min_charge_amount,
            min_charge_currency=min_charge_currency,
            rate_date_str=rate_date_str,
            rate_phrase=rate_phrase,
            rate_extra=rate_extra,
            subtotal_round=subtotal_round,
        )

    # Project 시트 (요약) — account 모드는 단일 Invoice 참조, per_project 모드는
    # 프로젝트별 Invoice 시트를 per_proj_invoice_meta 로 참조.
    # include_project_sheet=False 면 Project 시트 자체를 생성하지 않는다
    # (요약 시트를 사용하지 않는 계정용).
    if proj_results and include_project_sheet:
        write_project_sheet(
            wb, proj_results, company_name, billing_month,
            exchange_rate, margin_rate, invoice_date, bank_name,
            line_items=first_line_items,
            invoice_sku_rows=(None if is_per_project else first_sku_rows),
            invoice_rate_row=(None if is_per_project else first_rate_row),
            per_proj_invoice_meta=(per_proj_invoice_meta if is_per_project else None),
            currency=currency,
            min_charge_amount=min_charge_amount,
            min_charge_currency=min_charge_currency,
        )

    # GMP Price List 시트
    if price_list_file is not None:
        _copy_price_list_sheet(wb, price_list_file)

    if output_path:
        wb.save(str(output_path))
        return None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Invoice 시트 한 장 작성 (account 모드: 1회 / per_project 모드: 프로젝트마다)
# ─────────────────────────────────────────────────────────────────────────────
def _write_invoice_sheet(
    ws, line_items, company_name, billing_month,
    invoice_date, exchange_rate, margin_rate, bank_name,
    *, sku_order, currency, project_name,
    min_charge_amount: float = 500_000,
    min_charge_currency: str = "KRW",
    use_item_free_cap: bool = False,
    rate_date_str: str | None = None,
    rate_phrase: str = "최종 송금환율 기준",
    rate_extra: str = "",
    per_proj_ctx: dict | None = None,
    subtotal_round: int | None = None,
) -> tuple[list[dict], int | None, int | None]:
    """한 워크시트에 Invoice 레이아웃을 작성하고 (sku_rows, rate_row, sum_row) 반환.

    per_project 모드에서도 각 시트는 단일 프로젝트의 line_items 를 다루므로,
    G 열(수량) 은 account 모드와 동일한 waterfall 수식을 사용한다.
    (하드코딩 `_pp_tier_qty` 경로는 master 미등록 SKU 에서 0 이 되는 문제가
     있어 per_project 분리 시트 구조에선 불필요.)

    sum_row 는 per_project 모드에서 이 시트의 "합계(KRW)" 행 번호. 마지막
    시트의 총합 "청구금액(KRW)" 수식을 구성할 때 사용. account 모드는 None.
    """
    _n_items       = len([it for it in line_items if not _is_tax_sku(it.sku_name)])
    _estimated_max = 6 + 6 * _n_items + 3 + 3 + 2
    _set_white_background(ws, max_row=_estimated_max)
    _set_column_widths(ws)
    _write_image_area(ws)
    _write_invoice_info(ws, company_name, billing_month, invoice_date,
                        project_name=project_name)
    _write_table_header(ws)
    last_data_row, sku_rows = _write_data_rows(
        ws, line_items, sku_order=sku_order, currency=currency,
        billing_mode="account",
        use_item_free_cap=use_item_free_cap,
        subtotal_round=subtotal_round,
    )
    bottom_row, rate_row, sum_row = _write_summary_rows(
        ws, sku_rows, exchange_rate, margin_rate,
        last_data_row, billing_month, bank_name,
        currency=currency, line_items=line_items,
        min_charge_amount=min_charge_amount,
        min_charge_currency=min_charge_currency,
        rate_date_str=rate_date_str,
        rate_phrase=rate_phrase,
        rate_extra=rate_extra,
        per_proj_ctx=per_proj_ctx,
    )
    _write_vat_note(ws, bottom_row)
    _write_bottom_image(ws, bottom_row + 2)
    _set_freeze_pane(ws)
    return sku_rows, rate_row, sum_row


_INVALID_SHEET_CHARS = set(r':\/?*[]')

def _safe_sheet_title(name: str, used: list[str]) -> str:
    """Excel 시트명 제약(31자, 금지문자, 중복) 준수."""
    t = "".join(c for c in (name or "") if c not in _INVALID_SHEET_CHARS).strip()
    if not t:
        t = "Invoice"
    t = t[:31]
    base = t
    i = 2
    while t in used:
        suffix = f" ({i})"
        t = (base[: 31 - len(suffix)]) + suffix
        i += 1
    return t


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: 시트 전체 흰색 배경 — 실제 콘텐츠 범위만 채워야 Excel UsedRange
# 가 부풀려지지 않아 PDF 출력 시 빈 페이지/여백이 생기지 않는다.
# ─────────────────────────────────────────────────────────────────────────────
def _set_white_background(ws, max_row: int = 60, max_col: int = 12) -> None:
    ws.sheet_view.showGridLines = False
    white = _fill(C_WHITE)
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col):
        for cell in row:
            cell.fill = white


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: 열 너비
# ─────────────────────────────────────────────────────────────────────────────
def _set_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 32
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 18


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: Row 1 — 이미지 영역
# ─────────────────────────────────────────────────────────────────────────────
def _write_image_area(ws) -> None:
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    tt1_path = ASSETS_DIR / "tt1.png"
    tt2_path = ASSETS_DIR / "tt2.png"
    TOP_OFFSET_PX = 12

    if _HAS_IMAGE and tt1_path.exists():
        img1 = XLImage(str(tt1_path))
        orig_w, orig_h = img1.width, img1.height
        target_w = 400
        img1_h = int(orig_h * target_w / orig_w) if orig_w else target_w
        img1.width  = target_w
        img1.height = img1_h
        ws.row_dimensions[1].height = int(img1_h * 0.75) + 10
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 8
        ws.row_dimensions[4].height = 8
        marker1 = AnchorMarker(col=1, colOff=0, row=0,
                               rowOff=pixels_to_EMU(TOP_OFFSET_PX))
        size1 = XDRPositiveSize2D(pixels_to_EMU(target_w), pixels_to_EMU(img1_h))
        img1.anchor = OneCellAnchor(_from=marker1, ext=size1)
        ws.add_image(img1)
    else:
        ws.row_dimensions[1].height = 110
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 8
        ws.row_dimensions[4].height = 8
        ws["B1"] = "[tt1: SPH 회사 로고]"
        ws["B1"].font      = _font(bold=True, size=9, color="888888")
        ws["B1"].alignment = _align("left", "center")

    if _HAS_IMAGE and tt2_path.exists():
        img2 = XLImage(str(tt2_path))
        orig_w, orig_h = img2.width, img2.height
        target_w2 = 150
        img2_h = int(orig_h * target_w2 / orig_w) if orig_w else target_w2
        img2.width  = target_w2
        img2.height = img2_h
        marker2 = AnchorMarker(col=8, colOff=0, row=0,
                               rowOff=pixels_to_EMU(TOP_OFFSET_PX))
        size2 = XDRPositiveSize2D(pixels_to_EMU(target_w2), pixels_to_EMU(img2_h))
        img2.anchor = OneCellAnchor(_from=marker2, ext=size2)
        ws.add_image(img2)
    else:
        ws["I1"] = "[tt2: Google Maps Platform Invoice]"
        ws["I1"].font      = _font(bold=True, size=9, color="888888")
        ws["I1"].alignment = _align("center", "center")


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: Row 3-5 — 인보이스 메타데이터
# ─────────────────────────────────────────────────────────────────────────────
def _write_invoice_info(ws, company_name: str, billing_month: str,
                        invoice_date: str | None,
                        project_name: str = "") -> None:
    if invoice_date is None:
        invoice_date = date.today().strftime("%Y-%m-%d")

    year, month = int(billing_month[:4]), int(billing_month[5:7])
    last_day   = calendar.monthrange(year, month)[1]
    term_start = f"{billing_month}-01"
    term_end   = f"{billing_month}-{last_day:02d}"

    # Project Name 이 비어 있으면 해당 행(7) 자체를 숨긴다 — 레이아웃 상 빈 라벨
    # 한 줄이 남는 걸 피하기 위함.
    _has_proj = bool((project_name or "").strip())
    info_rows = [
        (5, "Invoice Date",         f":  {invoice_date}"),
        (6, "Billing Account Name", f":  {company_name}"),
    ]
    if _has_proj:
        info_rows.append((7, "Project Name", f":  {project_name}"))
    info_rows.append((8, "Term of Use", f":  {term_start}  ~  {term_end}"))
    if not _has_proj:
        ws.row_dimensions[7].hidden = True
    _no_border = Border()

    for row_num, label, value in info_rows:
        ws.row_dimensions[row_num].height = 20

        cell_b = ws.cell(row=row_num, column=C1, value=label)
        cell_b.font      = _font(bold=False, size=10)
        cell_b.fill      = _fill(C_WHITE)
        cell_b.alignment = _align("left", "center")
        cell_b.border    = _no_border

        cell_c = ws.cell(row=row_num, column=C1 + 1, value=value)
        cell_c.font      = _font(bold=False, size=10)
        cell_c.fill      = _fill(C_WHITE)
        cell_c.alignment = _align("left", "center")
        cell_c.border    = _no_border

        for c in range(C1 + 2, CN + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = _fill(C_WHITE)
            cell.border = _no_border


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: Row 6 — 테이블 헤더
# ─────────────────────────────────────────────────────────────────────────────
# 사용자 레퍼런스 이미지 기준 레이아웃:
#   Row 1-2 : 이미지
#   Row 3-4 : 빈 스페이서
#   Row 5   : Invoice Date
#   Row 6   : Billing Account Name
#   Row 7   : Project Name
#   Row 8   : Term of Use
#   Row 9   : 빈 스페이서
#   Row 10  : 테이블 헤더
#   Row 11+ : 데이터 (첫 SKU header_row=11, subtotal_row=16)
TABLE_HEADER_ROW = 10
HEADERS = ["API", "Usage", "Free Usage", "Subtotal",
           "할인 구간", "수량", "단가", "Amount"]

def _write_table_header(ws) -> None:
    ws.row_dimensions[TABLE_HEADER_ROW].height = 22
    for idx, label in enumerate(HEADERS):
        col  = C1 + idx
        cell = ws.cell(row=TABLE_HEADER_ROW, column=col, value=label)
        cell.font      = _font(bold=True, color=C_WHITE, size=10)
        cell.fill      = _fill(C_DARK)
        cell.alignment = _align("center", "center")
        cell.border    = _BORDER_ALL


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: Row 7+ — 데이터 행  (수식 기반 — GMP Price List 시트 참조)
# ─────────────────────────────────────────────────────────────────────────────
TIER_LABELS_ORDER = ["0-100K", "~500K", "~1M", "~5M", "~10M"]
PRICE_LIST_COLS   = ["D", "E", "F", "G", "H"]   # GMP Price List 시트의 tier 단가 컬럼
DATA_START_ROW    = TABLE_HEADER_ROW + 1        # = 7

# tax / VAT SKU는 인보이스 본문에서 제외 (별도 라인으로 표시되지 않음)
import re as _re
# 세금/VAT 판별 — 한글은 substring, 영문은 단어 경계 (elevation 의 'vat' 오탐 방지).
_TAX_RE = _re.compile(r"세금|\btax\b|\bvat\b", _re.IGNORECASE)
def _is_tax_sku(name: str) -> bool:
    return bool(_TAX_RE.search(name or ""))


def _col(idx: int) -> int:
    return C1 + idx


def _write_data_rows(ws, line_items: list,
                     sku_order: list[str] | None = None,
                     currency: str = "USD",
                     billing_mode: str = "account",
                     use_item_free_cap: bool = False,
                     subtotal_round: int | None = None) -> tuple[int, list[dict]]:
    """각 SKU마다 5개 tier 행 + 1개 소계 행을 수식 기반으로 작성한다.

    sku_order가 주어지면 그 순서대로 정렬; 목록에 없는 SKU는 뒤에 알파벳순.
    sku_order가 None이면 기본 알파벳순.

    currency='USD' (기본) : 단가/금액/소계 = "$#,##0.00" / "$#,##0"
    currency='KRW'        : 단가/금액/소계 = "₩#,##0"

    반환: (마지막 데이터 행 번호, sku_rows)
      sku_rows = [{"sku_name", "header_row", "subtotal_row"}, ...]
      — Project 시트에서 Invoice 셀 참조할 때 사용.
    """
    # 통화별 number format / 소계 반올림 자리수.
    # 사용자 레퍼런스 이미지 기준 USD 모드는 **전부 2 자리 소수** 표시.
    #   tier Amount  :  $700.00 / $2,240.00 / $325.26 …
    #   소계(Amount) :  $3,265.26
    # 따라서 소계도 ROUND(SUM, 2) 로 2 자리 저장 + 2 자리 포맷.
    # 총합(합계/청구금액) 쪽도 ROUND(, 2) 로 통일 — 이중 반올림 오차 제거.
    is_krw = (currency == "KRW")
    _tier_price_fmt  = '"\u20a9"#,##0'  if is_krw else '"$"#,##0.00'
    _tier_amount_fmt = '"\u20a9"#,##0'  if is_krw else '"$"#,##0.00'
    _subtotal_fmt    = '"\u20a9"#,##0'  if is_krw else '"$"#,##0.00'
    # ROUND(,N) 자리수. 외부에서 0/2 지정 가능. 미지정 시 통화별 기본.
    if subtotal_round in (0, 2):
        _subtotal_round = subtotal_round
    else:
        _subtotal_round = 0 if is_krw else 2
    # round=0 일 때 USD 표시에서 '.00' 제거 (KRW는 원래 무소수).
    # 소계뿐 아니라 구간별(tier) 단가·금액도 동일 자리수로 표시.
    # number_format 만 변경 — 셀 수식/저장값은 그대로라 결과값에 영향 없음.
    if _subtotal_round == 0:
        _subtotal_fmt    = '"₩"#,##0' if is_krw else '"$"#,##0'
        _tier_price_fmt  = '"₩"#,##0' if is_krw else '"$"#,##0'
        _tier_amount_fmt = '"₩"#,##0' if is_krw else '"$"#,##0'

    curr = DATA_START_ROW
    sku_rows: list[dict] = []

    # tax/VAT 제외
    filtered = [it for it in line_items if not _is_tax_sku(it.sku_name)]

    if sku_order:
        order_map = {name: idx for idx, name in enumerate(sku_order)}
        # 목록에 없는 것은 뒤에 붙임 (알파벳순)
        items = sorted(
            filtered,
            key=lambda x: (order_map.get(x.sku_name, len(order_map)), x.sku_name),
        )
    else:
        items = sorted(filtered, key=lambda x: x.sku_name)

    for item in items:
        header_row    = curr
        last_tier_row = curr + 4
        subtotal_row  = last_tier_row + 1

        b_ref = f"B{header_row}"
        e_ref = f"E{header_row}"

        # ── B(SKU명), C(사용량), D(무료-수식), E(소계-수식) — 5행 세로 병합 ─
        _merge_write(ws, header_row, last_tier_row, _col(0), item.sku_name,
                     fmt_data="center", wrap=True)
        _merge_write(ws, header_row, last_tier_row, _col(1), int(item.total_usage),
                     fmt_data="number")

        # Free Usage 수식.
        # - 기본 (account 모드): Price List SUMIF — Price List 의 full cap 반영.
        # - per_project (use_item_free_cap=True): rollover 로 프로젝트마다
        #   무료 배정량이 다를 수 있어 item.free_usage_cap 값을 직접 기록.
        if use_item_free_cap:
            free_formula = int(getattr(item, "free_usage_cap", 0) or 0)
        else:
            free_formula = (
                f"=SUMIF('GMP Price List'!$A:$A,{b_ref},'GMP Price List'!$C:$C)"
            )
        _merge_write(ws, header_row, last_tier_row, _col(2), free_formula,
                     fmt_data="center", num='"-"#,##0')

        subtotal_formula = f"=IF(C{header_row}-D{header_row}>0,C{header_row}-D{header_row},0)"
        _merge_write(ws, header_row, last_tier_row, _col(3), subtotal_formula,
                     fmt_data="number", num="#,##0")

        # ── 5개 tier 행: F(라벨), G(수량), H(단가), I(금액) ──
        # billing_mode:
        #   "account"     → G 는 waterfall 수식 (tier_limit 엄격 적용)
        #   "per_project" → G 는 Python 에서 계산된 tier 합산값을 하드코딩
        #                   (프로젝트별 독립 waterfall 의 tier 수량을 합쳐 놓은
        #                    값이라 tier_limit 을 초과할 수 있음 — 이게 정상)
        _is_per_proj = (billing_mode == "per_project")
        _pp_tier_qty: dict[int, int] = {}
        if _is_per_proj and getattr(item, "tier_breakdown", None):
            for _tb in item.tier_breakdown:
                _pp_tier_qty[int(_tb.tier_number)] = int(_tb.usage_in_tier)

        for i in range(5):
            r = header_row + i
            _cell_write(ws, r, _col(4), TIER_LABELS_ORDER[i], h="center")

            # G: tier별 수량 — per_project 모드면 하드코딩, 아니면 waterfall 수식
            if _is_per_proj:
                g_formula = _pp_tier_qty.get(i + 1, 0)
            elif i == 0:
                g_formula = f"=IF({e_ref}>100000,100000,{e_ref})"
            elif i == 1:
                g_formula = f"=IF({e_ref}>500000,400000,{e_ref}-G{header_row})"
            elif i == 2:
                g_formula = (
                    f"=IF({e_ref}>1000000,500000,"
                    f"{e_ref}-G{header_row}-G{header_row+1})"
                )
            elif i == 3:
                g_formula = (
                    f"=IF({e_ref}>5000000,4000000,"
                    f"{e_ref}-G{header_row}-G{header_row+1}-G{header_row+2})"
                )
            else:  # i == 4
                sub_expr = (
                    f"{e_ref}-G{header_row}-G{header_row+1}"
                    f"-G{header_row+2}-G{header_row+3}"
                )
                g_formula = f"=IF({sub_expr}>0,{sub_expr},0)"
            _cell_write(ws, r, _col(5), g_formula, h="right", num='#,##0;;"-"')

            # H: tier별 단가 (GMP Price List 해당 컬럼 참조)
            h_formula = (
                f"=SUMIF('GMP Price List'!$A:$A,{b_ref},"
                f"'GMP Price List'!${PRICE_LIST_COLS[i]}:${PRICE_LIST_COLS[i]})"
            )
            _cell_write(ws, r, _col(6), h_formula, h="right", num=_tier_price_fmt)

            # I: 수량 × 단가 / 1000
            i_formula = f"=G{r}*H{r}/1000"
            _cell_write(ws, r, _col(7), i_formula, h="right", num=_tier_amount_fmt)

        # ── 소계 행 (B:F 병합, G=SUM, I=ROUND(SUM,2)) ────────────────────────
        curr = subtotal_row

        for c in range(_col(0), _col(4) + 1):
            ws.cell(row=curr, column=c).fill   = _fill(C_SUB)
            ws.cell(row=curr, column=c).border = _BORDER_ALL

        ws.merge_cells(
            start_row=curr, start_column=_col(0),
            end_row=curr,   end_column=_col(4),
        )
        cell = ws.cell(row=curr, column=_col(0), value="소계")
        cell.font      = _font(bold=True)
        cell.fill      = _fill(C_SUB)
        cell.alignment = _align("center", "center")
        cell.border    = _BORDER_ALL

        _cell_write(ws, curr, _col(5),
                    f"=SUM(G{header_row}:G{last_tier_row})",
                    h="right", num='#,##0;;"-"', bold=True, bg=C_SUB)
        _cell_write(ws, curr, _col(6), "", h="center", bold=True, bg=C_SUB)
        _cell_write(ws, curr, _col(7),
                    f"=ROUND(SUM(I{header_row}:I{last_tier_row}),{_subtotal_round})",
                    h="right", num=_subtotal_fmt, bold=True, bg=C_SUB)

        sku_rows.append({
            "sku_name":     item.sku_name,
            "header_row":   header_row,
            "subtotal_row": subtotal_row,
        })
        curr = subtotal_row + 1

    return curr - 1, sku_rows


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: 합계 / 환율 / 청구금액 행
# ─────────────────────────────────────────────────────────────────────────────
MIN_CHARGE_KRW = 500_000   # Google Maps Platform 월 최소 사용비용 (legacy 기본값)


def _resolve_min_charge_krw(
    min_charge_amount: float,
    min_charge_currency: str,
    exchange_rate,
) -> int:
    """사용자 설정(금액·통화) 을 KRW 로 환산해 정수 반환.

    최소비용은 Invoice 에서 항상 '월최소사용비용(KRW)' 행으로 표시하므로,
    KRW 는 그대로, USD 는 환율을 곱해 정수 KRW 로 변환한다.
    amount ≤ 0 이면 적용하지 않는다는 의미로 0 반환.
    """
    try:
        amt = float(min_charge_amount)
    except (TypeError, ValueError):
        return 0
    if amt <= 0:
        return 0
    if min_charge_currency == "USD":
        try:
            rate = float(exchange_rate)
        except (TypeError, ValueError):
            rate = 0.0
        return int(round(amt * rate))
    return int(round(amt))


def _predict_total_krw(line_items, exchange_rate, margin_rate, is_krw: bool) -> int:
    """line_items 기반으로 최종 청구금액(KRW) 을 Python 에서 예측.

    500,000 미만인지 판정해 행 삽입을 결정하기 위한 값이며,
    엑셀 수식(정수 소계 합 × 환율 × 마진) 과 동일한 경로로 계산한다.
    """
    if not line_items:
        return 0
    _m = float(margin_rate) if margin_rate is not None else 1.0
    valid = [x for x in line_items if int(getattr(x, "total_usage", 0) or 0) > 0]
    # 각 SKU 소계는 엑셀에서 ROUND(SUM(tiers), 0) 로 정수 저장 → 동일 재현.
    int_subtotals = [
        int(Decimal(str(getattr(x, "subtotal_usd", 0) or 0))
            .quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        for x in valid
    ]
    total = sum(int_subtotals)
    if is_krw:
        # KRW 모드: 소계가 이미 원화. 환율 곱 없음.
        return round(total * _m)
    else:
        return round(total * float(exchange_rate) * _m)


def _render_summary_row(ws, r: int, label: str, value, *,
                        dark: bool, label_align: str = "left",
                        num_format: str = '"\u20a9"#,##0',
                        bold: bool = True) -> None:
    """합계 블록의 한 행을 병합+스타일 지정으로 한 번에 렌더."""
    bg   = C_DARK if dark else C_WHITE
    fg   = C_WHITE if dark else None
    for c in range(_col(0), _col(6) + 1):
        ws.cell(row=r, column=c).fill   = _fill(bg)
        ws.cell(row=r, column=c).border = _BORDER_ALL
    ws.merge_cells(start_row=r, start_column=_col(0),
                   end_row=r,   end_column=_col(6))
    cell = ws.cell(row=r, column=_col(0), value=label)
    cell.font      = _font(bold=bold, color=fg) if fg else _font(bold=bold)
    cell.fill      = _fill(bg)
    cell.alignment = _align(label_align, "center")
    cell.border    = _BORDER_ALL

    cell8 = ws.cell(row=r, column=_col(7), value=value)
    cell8.font          = _font(bold=bold, color=fg) if fg else _font(bold=bold)
    cell8.fill          = _fill(bg)
    cell8.alignment     = _align("right", "center")
    cell8.number_format = num_format
    cell8.border        = _BORDER_ALL


def _write_summary_rows(ws, sku_rows: list[dict], exchange_rate, margin_rate,
                        last_data_row: int,
                        billing_month: str = "2026-01",
                        bank_name: str = "하나은행",
                        currency: str = "USD",
                        line_items=None,
                        min_charge_amount: float = 500_000,
                        min_charge_currency: str = "KRW",
                        rate_date_str: str | None = None,
                        rate_phrase: str = "최종 송금환율 기준",
                        rate_extra: str = "",
                        per_proj_ctx: dict | None = None,
                        ) -> tuple[int, int | None, int | None]:
    """합계 / 환율 / 청구금액(KRW) 행 작성. 최소 사용비용 룰 적용.

    min_charge_amount/min_charge_currency: 계정별 최소사용비용 설정.
      KRW 면 그대로, USD 면 환율로 환산해 KRW 로 비교·표시한다.
      0 이면 최소비용 룰 자체를 적용하지 않는다(회사 정책에 따라 면제 가능).

    - 예상 청구금액(KRW) 이 최소사용비용(KRW) 미만이면 청구금액 위에
      "월최소사용비용(KRW)" 행을 삽입하고 청구금액 = 최소비용 으로 보정.
    - USD 모드: 합계(USD) / 환율 / [합계(KRW) / 월최소사용비용(KRW)]* /
      청구금액(KRW)
    - KRW 모드: [합계(KRW) / 월최소사용비용(KRW)]* / 청구금액(KRW)

    per_proj_ctx (per_project 모드에서만 전달):
      {"is_last": bool, "prev_sum_refs": list[str], "self_sheet_title": str}
      - 각 프로젝트 시트의 "청구금액(KRW)" 라벨은 "합계(KRW)" 로 바뀐다
        (해당 시트 단위의 합계라는 의미).
      - is_last=True 이면 그 "합계(KRW)" 아래에 실제 "청구금액(KRW)" 행을 추가해
        모든 프로젝트 시트의 합계 셀을 SUM 한다.

    반환: (마지막 행 번호, 환율 행 번호 | None, 이 시트의 합계 KRW 행 번호 | None)
    """
    is_per_proj = per_proj_ctx is not None
    is_last_proj = bool(is_per_proj and per_proj_ctx.get("is_last"))
    # per_project 모드에서는 "청 구 금 액(KRW)" 라벨을 "합        계(KRW)" 로 교체.
    _final_krw_label = "합        계(KRW)" if is_per_proj else "청 구 금 액(KRW)"
    # per_project 모드 라벨은 상단 합계(USD)/환율과 같은 좌측 정렬. account 모드는 기존 중앙 유지.
    _final_krw_align = "left" if is_per_proj else "center"
    is_krw = (currency == "KRW")
    _m = float(margin_rate) if margin_rate is not None else 1.0

    year, month = int(billing_month[:4]), int(billing_month[5:7])
    last_day     = calendar.monthrange(year, month)[1]
    # rate_date_str(YYYY.MM.DD) 이 주어지면 우선 사용, 아니면 billing_month
    # 마지막 날로 자동 생성 (legacy 동작 유지).
    last_day_str = rate_date_str or f"{year}.{month:02d}.{last_day:02d}"
    # 환율 라벨의 고정문구 + 추가문구 조합.
    _rate_note = (rate_phrase or "최종 송금환율 기준").strip()
    _rate_extra = (rate_extra or "").strip()
    if _rate_extra:
        _rate_note = f"{_rate_note} {_rate_extra}"

    sub_refs = ",".join(f"I{s['subtotal_row']}" for s in sku_rows) if sku_rows else ""

    # ── 최소사용비용 적용 여부 판정 ────────────────────────────────────────
    # 규칙: 계정에 **사용량이 있는 달** 이면 최소비용 적용.
    #   - 예상 청구금액이 최소비용 미만이면 최소비용 행 삽입하고 청구액 보정.
    #   - 예상 청구금액 = 0 인 경우(전액 무료 제공 범위 내) 도 포함. Google 실제
    #     정책 상 사용 있는 월은 최소 ₩500,000 부과 대상.
    #   - 사용량 자체가 0 인 달(유령 계정) 은 청구서를 내지 않으므로 제외.
    min_charge_krw = _resolve_min_charge_krw(
        min_charge_amount, min_charge_currency, exchange_rate,
    )
    predicted_krw = _predict_total_krw(line_items, exchange_rate, margin_rate, is_krw)
    _has_any_usage = any(
        int(getattr(_it, "total_usage", 0) or 0) > 0
        for _it in (line_items or [])
    )
    min_charge_applied = (
        min_charge_krw > 0
        and _has_any_usage
        and predicted_krw < min_charge_krw
    )

    r = last_data_row + 1

    def _append_grand_total(r_cur: int, self_sum_row: int) -> int:
        """per_project 마지막 시트에만: '합계(KRW)' 아래에 모든 시트 합산 '청구금액(KRW)' 행."""
        if not is_last_proj:
            return r_cur
        prev_refs = list((per_proj_ctx or {}).get("prev_sum_refs") or [])
        self_title = (per_proj_ctx or {}).get("self_sheet_title") or ""
        if self_title:
            safe = self_title.replace("'", "''")
            self_ref = f"'{safe}'!I{self_sum_row}"
        else:
            self_ref = f"I{self_sum_row}"
        all_refs = prev_refs + [self_ref]
        gt_formula = f"=SUM({','.join(all_refs)})" if all_refs else 0
        r_cur += 1
        _render_summary_row(ws, r_cur, "청 구 금 액(KRW)", gt_formula,
                            dark=True, label_align="left")
        return r_cur

    # ══════════════════════════════════════════════════════════════════════
    # KRW 모드
    # ══════════════════════════════════════════════════════════════════════
    if is_krw:
        if not min_charge_applied:
            # 기존 1행 구조: (per_proj 모드는 "합계(KRW)") = SUM(sub_refs) × margin
            if sub_refs:
                formula = (f"=ROUND(SUM({sub_refs}),2)" if _m == 1.0
                           else f"=ROUND(SUM({sub_refs})*{_m},2)")
            else:
                formula = 0
            _render_summary_row(ws, r, _final_krw_label, formula,
                                dark=True, label_align=_final_krw_align)
            _sum_row = r if is_per_proj else None
            if is_last_proj:
                r = _append_grand_total(r, r)
            return r, None, _sum_row

        # 최소사용비용 적용: 3행 (합계(KRW) / 월최소사용비용(KRW) / 청구금액(KRW))
        subtotal_formula = (f"=ROUND(SUM({sub_refs}),2)" if _m == 1.0
                            else f"=ROUND(SUM({sub_refs})*{_m},2)") if sub_refs else 0
        _render_summary_row(ws, r, "합        계(KRW)", subtotal_formula, dark=False)
        r += 1
        _render_summary_row(ws, r, "월최소사용비용(KRW)", min_charge_krw, dark=False)
        r += 1
        _render_summary_row(ws, r, _final_krw_label, min_charge_krw,
                            dark=True, label_align=_final_krw_align)
        _sum_row = r if is_per_proj else None
        if is_last_proj:
            r = _append_grand_total(r, r)
        return r, None, _sum_row

    # ══════════════════════════════════════════════════════════════════════
    # USD 모드
    # ══════════════════════════════════════════════════════════════════════
    # 합계(USD) — 2 자리 소수로 저장(표시). 원화 환산(I{usd}*I{rate}) 시 이 값을
    # 그대로 참조하므로 이중 반올림 오차 없음.
    usd_row = r
    _render_summary_row(ws, r, "합        계(USD)",
                        (f"=ROUND(SUM({sub_refs}),2)" if sub_refs else 0),
                        dark=False, num_format='"$"#,##0.00')
    # 환율
    r += 1
    rate_row = r
    _render_summary_row(
        ws, r,
        f"환        율({bank_name} {last_day_str} {_rate_note})",
        float(exchange_rate),
        dark=False, num_format='"\u20a9"#,##0.00',
    )

    if not min_charge_applied:
        # 기존 3행 구조: (per_proj 모드는 "합계(KRW)") = ROUND(I{usd}*I{rate}*margin, 2)
        r += 1
        formula = (f"=ROUND(I{usd_row}*I{rate_row},2)" if _m == 1.0
                   else f"=ROUND(I{usd_row}*I{rate_row}*{_m},2)")
        _render_summary_row(ws, r, _final_krw_label, formula,
                            dark=True, label_align=_final_krw_align)
        _sum_row = r if is_per_proj else None
        if is_last_proj:
            r = _append_grand_total(r, r)
        return r, rate_row, _sum_row

    # 최소사용비용 적용: 합계(KRW) + 월최소사용비용(KRW) + 청구금액(KRW)
    r += 1
    subtotal_krw_formula = (f"=ROUND(I{usd_row}*I{rate_row},2)" if _m == 1.0
                            else f"=ROUND(I{usd_row}*I{rate_row}*{_m},2)")
    _render_summary_row(ws, r, "합        계(KRW)", subtotal_krw_formula, dark=False)
    r += 1
    _render_summary_row(ws, r, "월최소사용비용(KRW)", MIN_CHARGE_KRW, dark=False)
    r += 1
    _render_summary_row(ws, r, _final_krw_label, MIN_CHARGE_KRW,
                        dark=True, label_align=_final_krw_align)
    _sum_row = r if is_per_proj else None
    if is_last_proj:
        r = _append_grand_total(r, r)
    return r, rate_row, _sum_row


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: (부가세 별도)
# ─────────────────────────────────────────────────────────────────────────────
def _write_vat_note(ws, krw_row: int) -> None:
    r = krw_row + 1
    ws.row_dimensions[r].height = 16

    # 병합 전 비앵커 셀 스타일 먼저
    for c in range(C1 + 1, CN + 1):
        ws.cell(row=r, column=c).fill   = _fill(C_WHITE)
        ws.cell(row=r, column=c).border = Border()

    ws.merge_cells(start_row=r, start_column=C1, end_row=r, end_column=CN)
    cell = ws.cell(row=r, column=C1, value="(부가세 별도)")
    cell.font      = _font(size=9, color="595959")
    cell.fill      = _fill(C_WHITE)
    cell.alignment = _align("right", "center")
    cell.border    = Border()


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: 하단 tt3 이미지
# ─────────────────────────────────────────────────────────────────────────────
def _write_bottom_image(ws, start_row: int) -> None:
    tt3_path = ASSETS_DIR / "tt3.png"

    if _HAS_IMAGE and tt3_path.exists():
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        img3 = XLImage(str(tt3_path))
        orig_w, orig_h = img3.width, img3.height
        target_w = 400
        img3_h = int(orig_h * target_w / orig_w) if orig_w else target_w
        img3.width  = target_w
        img3.height = img3_h
        ws.row_dimensions[start_row].height = int(img3_h * 0.75) + 10

        col_widths_px   = [10*7, 32*7, 14*7, 14*7, 14*7, 14*7, 14*7, 14*7, 18*7]
        table_right_px  = sum(col_widths_px)
        img_left_px     = table_right_px - target_w

        anchor_col        = 0
        anchor_col_off_px = 0
        cumul = 0
        for i, w in enumerate(col_widths_px):
            if cumul + w > img_left_px:
                anchor_col        = i
                anchor_col_off_px = img_left_px - cumul
                break
            cumul += w

        marker = AnchorMarker(
            col=anchor_col,
            colOff=pixels_to_EMU(anchor_col_off_px),
            row=start_row - 1,
            rowOff=0,
        )
        size = XDRPositiveSize2D(pixels_to_EMU(target_w), pixels_to_EMU(img3_h))
        img3.anchor = OneCellAnchor(_from=marker, ext=size)
        ws.add_image(img3)
    else:
        ws.row_dimensions[start_row].height = 80

        # 병합 전 비앵커 셀 스타일 먼저
        for c in range(C1 + 1, CN + 1):
            ws.cell(row=start_row, column=c).fill   = _fill(C_WHITE)
            ws.cell(row=start_row, column=c).border = Border()

        ws.merge_cells(
            start_row=start_row, start_column=C1,
            end_row=start_row,   end_column=CN,
        )
        cell = ws.cell(row=start_row, column=C1,
                       value="[tt3: Google Cloud Premier Partner 배지]")
        cell.font      = _font(size=9, color="888888")
        cell.fill      = _fill(C_WHITE)
        cell.alignment = _align("right", "center")
        cell.border    = Border()


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: 틀 고정 해제
# ─────────────────────────────────────────────────────────────────────────────
def _set_freeze_pane(ws) -> None:
    ws.freeze_panes = None


# ─────────────────────────────────────────────────────────────────────────────
# 저수준 셀 작성 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _cell_write(ws, row: int, col: int, value,
                h="center", v="center", wrap=False,
                num="General", bold=False,
                color="000000", bg=C_WHITE) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font          = _font(bold=bold, color=color)
    cell.fill          = _fill(bg)
    cell.alignment     = _align(h, v, wrap)
    cell.number_format = num
    cell.border        = _BORDER_ALL
    ws.row_dimensions[row].height = 18


def _merge_write(ws, r1: int, r2: int, col: int, value,
                 fmt_data="center", wrap=False,
                 bold=False, color="000000", bg=C_WHITE,
                 num: str | None = None, h: str | None = None) -> None:
    """단일 열 세로 병합 (r1~r2, col 고정).
    병합 전에 범위 내 모든 셀(r1~r2 전체)에 테두리를 먼저 입혀 왼쪽 선 유실 방지.
    num / h 를 명시하면 fmt_data 프리셋을 덮어쓴다."""
    _h   = "right" if fmt_data == "number" else "center"
    _num = "#,##0" if fmt_data == "number" else "General"
    if h   is not None: _h   = h
    if num is not None: _num = num
    h, num = _h, _num

    # ① 병합 전: r1 포함 전체 범위(r1~r2)에 스타일 먼저 적용
    for r in range(r1, r2 + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill   = _fill(bg)
        cell.font   = _font(bold=bold, color=color)
        cell.border = _BORDER_ALL

    # ② 병합 실행
    ws.merge_cells(start_row=r1, start_column=col,
                   end_row=r2,   end_column=col)

    # ③ 앵커 셀(r1)에 값·정렬 추가 — 테두리는 ①에서 확정됨
    cell = ws.cell(row=r1, column=col, value=value)
    cell.font          = _font(bold=bold, color=color)
    cell.fill          = _fill(bg)
    cell.alignment     = _align(h, "center", wrap)
    cell.number_format = num
    cell.border        = _BORDER_ALL


# ─────────────────────────────────────────────────────────────────────────────
# 내부 헬퍼: GMP Price List 시트 복제 (MergedCell 완전 방어)
# ─────────────────────────────────────────────────────────────────────────────
def _copy_price_list_sheet(wb: Workbook, price_list_file) -> None:
    """원본 엑셀 첫 번째 시트를 'GMP Price List' 탭으로 복제."""
    try:
        if hasattr(price_list_file, "read"):
            price_list_file.seek(0)
            raw = price_list_file.read()
            src_wb = load_workbook(io.BytesIO(raw), data_only=True)
        else:
            src_wb = load_workbook(str(price_list_file), data_only=True)
    except Exception:
        return

    src_ws = src_wb.worksheets[0]
    dst_ws = wb.create_sheet("GMP Price List")
    dst_ws.sheet_view.showGridLines = False

    # 탭 색상
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = copy(src_ws.sheet_properties.tabColor)

    from openpyxl.utils import get_column_letter

    # 실제 데이터 범위로 제한 (시트 dim A1:Z1005 같이 과대 범위 순회 방지)
    max_row = min(src_ws.max_row or 1, 200)
    max_col = min(src_ws.max_column or 1, 30)

    # 시트 기본값
    default_col_w = src_ws.sheet_format.defaultColWidth or 8.43
    if src_ws.sheet_format.defaultColWidth:
        dst_ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
    if src_ws.sheet_format.defaultRowHeight:
        dst_ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight

    # 명시적 열 너비 복사
    explicit_cols = set()
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_col = dst_ws.column_dimensions[col_letter]
        if col_dim.width and col_dim.width > 0:
            dst_col.width = col_dim.width * 1.2
            explicit_cols.add(col_letter)
        dst_col.hidden = col_dim.hidden

    # column_dimensions에 없는 열(기본 너비 사용) → 셀 내용 길이로 자동 산정
    auto_w: dict[str, int] = {}
    for row in src_ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            cl = get_column_letter(cell.column)
            if cl not in explicit_cols:
                auto_w[cl] = max(auto_w.get(cl, 0), len(str(cell.value)))
    for cl, w in auto_w.items():
        dst_ws.column_dimensions[cl].width = max((w + 4) * 1.3, default_col_w)

    # I열 추가 확장
    if dst_ws.column_dimensions["I"].width:
        dst_ws.column_dimensions["I"].width *= 1.25

    # 행 높이
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_row = dst_ws.row_dimensions[row_idx]
        dst_row.height = row_dim.height
        dst_row.hidden = row_dim.hidden

    # 병합 범위 선이식
    for merged_range in list(src_ws.merged_cells.ranges):
        try:
            dst_ws.merge_cells(str(merged_range))
        except Exception:
            pass

    # 셀 값 + 스타일 복사 — 실제 데이터 범위로 제한
    for row in src_ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
        for src_cell in row:
            if isinstance(src_cell, MergedCell):
                continue

            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)

            if isinstance(dst_cell, MergedCell):
                continue

            # A 열(SKU 이름) 은 trailing/leading 공백 정리. 사용자 업로드 Price
            # List 에 'Places API Place Details Enterprise ' 처럼 끝 공백이 있는
            # 행이 있어, Invoice 의 SUMIF(exact match) 가 빠지는 문제 방지.
            _v = src_cell.value
            if src_cell.column == 1 and isinstance(_v, str):
                _v = _v.strip()
            dst_cell.value = _v

            if src_cell.has_style:
                dst_cell.font          = copy(src_cell.font)
                dst_cell.fill          = copy(src_cell.fill)
                dst_cell.border        = copy(src_cell.border)
                dst_cell.alignment     = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                if src_cell.protection:
                    dst_cell.protection = copy(src_cell.protection)

    # "COST PER THOUSAND (CPM)" 셀 border 누락 보정 — 실제 데이터 범위만 순회
    _thin = Side(style="thin")
    _border_fix = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for row in dst_ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and "COST PER THOUSAND" in str(cell.value).upper():
                cell.border = _border_fix


# ─────────────────────────────────────────────────────────────────────────────
# CLI 진입점
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="t1 레이아웃 인보이스 생성기",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("input_file", nargs="?", default="billing.csv",
                        help="사용고지서 파일 (기본: billing.csv)")
    parser.add_argument("-m", "--billing-month", default="2026-03", metavar="YYYY-MM")
    parser.add_argument("-e", "--exchange-rate", type=float, default=1427.87, metavar="RATE")
    parser.add_argument("-r", "--margin-rate",   type=float, default=1.12,    metavar="RATE")
    parser.add_argument("-c", "--company",        default=None,
                        help="특정 회사만 필터링 (예: coupang)")
    parser.add_argument("-o", "--output",         default="invoice_output.xlsx", metavar="FILE")
    parser.add_argument("--invoice-date",         default=None,
                        help="인보이스 날짜 YYYY-MM-DD (기본: 오늘)")
    args = parser.parse_args()

    INPUT_FILE  = Path(args.input_file)
    OUTPUT_FILE = BASE_DIR / args.output
    EXCHANGE    = Decimal(str(args.exchange_rate))
    MARGIN      = Decimal(str(args.margin_rate))

    print(f"[1/3] 데이터 전처리: {INPUT_FILE.name}")
    raw_rows = preprocess_usage_file(
        INPUT_FILE, args.billing_month,
        company_filter=args.company,
    )

    print(f"[2/3] 과금 계산 ({len(raw_rows)} 행)")
    sku_master = load_sku_master(SKU_MASTER_ROWS)
    usage_rows = load_usage_rows(raw_rows)
    line_items = calculate_billing(usage_rows, sku_master, EXCHANGE, MARGIN)

    company_display = args.company or "All Companies"
    print(f"[3/3] 인보이스 생성: {OUTPUT_FILE.name}")
    generate_formatted_invoice(
        line_items    = line_items,
        company_name  = company_display,
        billing_month = args.billing_month,
        exchange_rate = EXCHANGE,
        margin_rate   = MARGIN,
        invoice_date  = args.invoice_date,
        output_path   = OUTPUT_FILE,
    )

    print(f"\n완료 -> {OUTPUT_FILE}")
    print(f"  항목 수: {len(line_items)}건")
    for it in line_items:
        print(f"  - {it.sku_name}: ${float(it.subtotal_usd):,.2f}")
